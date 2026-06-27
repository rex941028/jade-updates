import base64
import json
import os
import re
import shutil
import sqlite3
import sys
import threading
import tkinter as tk
import urllib.error
import urllib.request
from tkinter import filedialog, messagebox, ttk

import openpyxl
from tkcalendar import Calendar

# ── Constants ─────────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, 'data', 'customers.db')

# ── 版本與自動更新 ─────────────────────────────────────────────────────────────
# 每次推送更新時，同步修改此版本號。
APP_VERSION = "1.0.3"

# 將此 URL 設為你 GitHub 上 update.json 的 Raw 連結。
# 範例：https://raw.githubusercontent.com/你的帳號/jade-updates/main/update.json
# 留空則停用自動更新。
UPDATE_CHECK_URL = "https://raw.githubusercontent.com/rex941028/jade-updates/main/update.json"

JADE       = '#2D6A4F'
JADE_DARK  = '#1B4332'
JADE_MID   = '#D8F3E8'
JADE_LIGHT = '#52B788'
BG         = '#F5F7F5'
WHITE      = '#FFFFFF'
TEXT       = '#2C3E50'
GRAY       = '#6C757D'
RED        = '#C0392B'
BLUE       = '#1565C0'
FONT       = 'Microsoft JhengHei'

# ── Database ──────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS customers (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            shopee_account TEXT UNIQUE NOT NULL,
            real_name      TEXT DEFAULT '',
            line_account   TEXT DEFAULT '',
            birthday       TEXT DEFAULT '',
            preferences    TEXT DEFAULT '',
            notes          TEXT DEFAULT '',
            created_at     TEXT DEFAULT (datetime('now','localtime')),
            updated_at     TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS orders (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id         TEXT UNIQUE NOT NULL,
            shopee_account   TEXT NOT NULL,
            order_status     TEXT DEFAULT '',
            refund_status    TEXT DEFAULT '',
            order_date       TEXT DEFAULT '',
            product_name     TEXT DEFAULT '',
            product_option   TEXT DEFAULT '',
            quantity         INTEGER DEFAULT 1,
            product_price    REAL DEFAULT 0,
            paid_amount      REAL DEFAULT 0,
            original_price   REAL DEFAULT 0,
            seller_coupon    REAL DEFAULT 0,
            transaction_fee  REAL DEFAULT 0,
            service_fee      REAL DEFAULT 0,
            payment_fee      REAL DEFAULT 0,
            return_shipping  REAL DEFAULT 0,
            city             TEXT DEFAULT '',
            district         TEXT DEFAULT '',
            payment_method   TEXT DEFAULT '',
            shipping_method  TEXT DEFAULT '',
            tracking_number  TEXT DEFAULT '',
            order_complete_time TEXT DEFAULT '',
            buyer_note       TEXT DEFAULT '',
            seller_note      TEXT DEFAULT '',
            created_at       TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS import_logs (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            filename       TEXT,
            imported_at    TEXT DEFAULT (datetime('now','localtime')),
            new_orders     INTEGER DEFAULT 0,
            updated_orders INTEGER DEFAULT 0,
            new_customers  INTEGER DEFAULT 0
        );
    ''')
    for col_def in (
        "ALTER TABLE orders ADD COLUMN original_price   REAL DEFAULT 0",
        "ALTER TABLE orders ADD COLUMN seller_coupon    REAL DEFAULT 0",
        "ALTER TABLE orders ADD COLUMN transaction_fee  REAL DEFAULT 0",
        "ALTER TABLE orders ADD COLUMN service_fee      REAL DEFAULT 0",
        "ALTER TABLE orders ADD COLUMN payment_fee      REAL DEFAULT 0",
        "ALTER TABLE orders ADD COLUMN return_shipping  REAL DEFAULT 0",
        "ALTER TABLE orders ADD COLUMN buyer_note       TEXT DEFAULT ''",
        "ALTER TABLE orders ADD COLUMN seller_note      TEXT DEFAULT ''",
        "ALTER TABLE orders ADD COLUMN refunded_price   REAL DEFAULT 0",
    ):
        try:
            conn.execute(col_def)
        except Exception:
            pass
    conn.commit()
    conn.close()


# ── Excel helpers (file opened exactly once) ──────────────────────────────────

def _parse_excel(filepath):
    """
    Read Excel into memory, close file immediately.
    Multi-item orders: product names are joined with ' | ' in product_name field.
    Returns (order_map, price_acc, filename).
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    col_lookup = {}
    for i, h in enumerate(raw_headers):
        if h is not None:
            col_lookup[str(h).strip().replace('\n', '').replace(' ', '')] = i

    def idx(name):
        return col_lookup.get(name.replace('\n', '').replace(' ', ''))

    def gv(row, name, default=''):
        i = idx(name)
        if i is None or i >= len(row):
            return default
        v = row[i]
        return str(v).strip() if v is not None else default

    def gf(row, name):
        try:    return float(gv(row, name, '0') or 0)
        except: return 0.0

    def gi(row, name):
        try:    return int(float(gv(row, name, '1') or 1))
        except: return 1

    order_map          = {}
    price_acc          = {}
    refunded_price_acc = {}  # order_id → sum of refunded items' original price
    product_items      = {}  # order_id → item strings (refunded items prefixed with [退])
    qty_acc            = {}  # order_id → total quantity

    for row in ws.iter_rows(min_row=2, values_only=True):
        oid  = gv(row, '訂單編號')
        acct = gv(row, '買家帳號')
        if not oid or not acct:
            continue

        row_refund = gv(row, '退貨/退款狀態')
        item_price = gf(row, '商品原價')
        is_row_refunded = bool(row_refund.strip())

        # Accumulate prices (track refunded portion separately)
        price_acc[oid] = price_acc.get(oid, 0.0) + item_price
        if is_row_refunded:
            refunded_price_acc[oid] = refunded_price_acc.get(oid, 0.0) + item_price

        # Accumulate product item description (mark refunded items visually)
        name   = gv(row, '商品名稱')
        option = gv(row, '商品選項名稱')
        qty    = gi(row, '數量')
        item   = name
        if option:  item += f'（{option}）'
        if qty > 1: item += f' ×{qty}'
        if is_row_refunded:
            item = '[退] ' + item
        product_items.setdefault(oid, []).append(item)
        qty_acc[oid] = qty_acc.get(oid, 0) + qty

        # Order-level fields: first row sets the base; subsequent rows only patch
        row_rs = gf(row, '退貨運費')
        if oid not in order_map:
            order_map[oid] = dict(
                order_id=oid, shopee_account=acct,
                order_status=gv(row, '訂單狀態'),
                refund_status=row_refund,
                order_date=gv(row, '訂單成立日期'),
                product_name='',  # filled after loop
                product_option='',
                quantity=1,       # filled after loop
                product_price=gf(row, '商品總價'),
                paid_amount=gf(row, '買家總支付金額'),
                city=gv(row, '城市'), district=gv(row, '行政區'),
                payment_method=gv(row, '付款方式'),
                shipping_method=gv(row, '寄送方式'),
                tracking_number=gv(row, '包裹查詢號碼'),
                order_complete_time=gv(row, '訂單完成時間'),
                buyer_note=gv(row, '買家備註'),
                seller_note=gv(row, '備註'),
                seller_coupon=gf(row, '賣家負擔優惠券'),
                transaction_fee=gf(row, '成交手續費'),
                service_fee=gf(row, '其他服務費'),
                payment_fee=gf(row, '金流與系統處理費'),
                return_shipping=row_rs,
            )
        else:
            od = order_map[oid]
            # Preserve any non-empty refund status (don't overwrite with empty later rows)
            if is_row_refunded and not od['refund_status']:
                od['refund_status'] = row_refund
            # Take max return_shipping to avoid double-counting on fully-refunded orders
            if row_rs > (od['return_shipping'] or 0):
                od['return_shipping'] = row_rs
            # Use latest order_complete_time
            new_ct = gv(row, '訂單完成時間')
            if new_ct and (not od['order_complete_time'] or new_ct > od['order_complete_time']):
                od['order_complete_time'] = new_ct

    # Merge accumulated info back into order_map
    for oid in order_map:
        order_map[oid]['product_name']   = ' | '.join(product_items.get(oid, []))
        order_map[oid]['quantity']       = qty_acc.get(oid, 1)
        order_map[oid]['refunded_price'] = refunded_price_acc.get(oid, 0.0)

    wb.close()
    return order_map, price_acc, os.path.basename(filepath)


def _find_dup_orders(order_map):
    conn = get_db()
    dups = [oid for oid in order_map
            if conn.execute('SELECT 1 FROM orders WHERE order_id=?', (oid,)).fetchone()]
    conn.close()
    return dups


def _write_to_db(order_map, price_acc, filename, skip_existing=False):
    conn = get_db()
    new_c = new_o = upd_o = 0

    for oid, od in order_map.items():
        od['original_price'] = price_acc[oid]
        acct = od['shopee_account']

        if not conn.execute('SELECT 1 FROM customers WHERE shopee_account=?', (acct,)).fetchone():
            conn.execute('INSERT INTO customers (shopee_account) VALUES (?)', (acct,))
            new_c += 1

        existing = conn.execute(
            'SELECT order_complete_time FROM orders WHERE order_id=?', (oid,)
        ).fetchone()

        if existing:
            if not skip_existing:
                stored_ct = existing['order_complete_time'] or ''
                new_ct    = od['order_complete_time'] or ''
                is_stale  = (stored_ct and not new_ct) or \
                            (stored_ct and new_ct and new_ct < stored_ct)
                if not is_stale:
                    conn.execute(
                        'UPDATE orders SET '
                        'order_status=:order_status, refund_status=:refund_status, '
                        'order_date=:order_date, product_name=:product_name, '
                        'product_option=:product_option, quantity=:quantity, '
                        'product_price=:product_price, paid_amount=:paid_amount, '
                        'original_price=:original_price, seller_coupon=:seller_coupon, '
                        'transaction_fee=:transaction_fee, service_fee=:service_fee, '
                        'payment_fee=:payment_fee, return_shipping=:return_shipping, '
                        'refunded_price=:refunded_price, '
                        'city=:city, district=:district, payment_method=:payment_method, '
                        'shipping_method=:shipping_method, tracking_number=:tracking_number, '
                        'order_complete_time=:order_complete_time, '
                        'buyer_note=:buyer_note, seller_note=:seller_note '
                        'WHERE order_id=:order_id', od)
            upd_o += 1
        else:
            conn.execute(
                'INSERT INTO orders ('
                'order_id, shopee_account, order_status, refund_status, order_date, '
                'product_name, product_option, quantity, product_price, paid_amount, '
                'original_price, seller_coupon, transaction_fee, service_fee, '
                'payment_fee, return_shipping, refunded_price, city, district, payment_method, '
                'shipping_method, tracking_number, order_complete_time, '
                'buyer_note, seller_note) VALUES ('
                ':order_id, :shopee_account, :order_status, :refund_status, :order_date, '
                ':product_name, :product_option, :quantity, :product_price, :paid_amount, '
                ':original_price, :seller_coupon, :transaction_fee, :service_fee, '
                ':payment_fee, :return_shipping, :refunded_price, :city, :district, :payment_method, '
                ':shipping_method, :tracking_number, :order_complete_time, '
                ':buyer_note, :seller_note)', od)
            new_o += 1

    conn.execute(
        'INSERT INTO import_logs (filename,new_orders,updated_orders,new_customers) VALUES (?,?,?,?)',
        (filename, new_o, upd_o, new_c))
    conn.commit()
    conn.close()
    return new_c, new_o, upd_o


# ── Notes → LINE scanner ─────────────────────────────────────────────────────

# Matches "LINE會員_XXX" where XXX runs to 。 / ， / , / newline / end-of-string.
# Handles case variants (line/LINE) and full-width Ｎ used in some notes.
_LINE_EXTRACT_RE = re.compile(r'li[nｎＮ]e會員_([^。，,\n]+)', re.IGNORECASE)
# Reject candidates that look like product/action notes rather than LINE account names.
_NOTE_ACTION_RE  = re.compile(r'有送|此單|使用|記得|下次|已送')


def _extract_line_candidate(note: str):
    """Return the LINE account embedded in a seller_note, or None."""
    m = _LINE_EXTRACT_RE.search(note)
    if not m:
        return None
    candidate = m.group(1).strip()
    # Trim trailing parenthetical remarks that look like non-name annotations
    # e.g. "陳隆盛 0201加入會員" → "陳隆盛"
    # Strategy: if space exists and text after space starts with digit, trim there
    sp = candidate.find(' ')
    if sp != -1 and sp < len(candidate) - 1 and candidate[sp + 1].isdigit():
        candidate = candidate[:sp].strip()
    if not candidate or len(candidate) > 35:
        return None
    if _NOTE_ACTION_RE.search(candidate):
        return None
    return candidate


def _scan_line_from_notes():
    """Scan all seller_notes for LINE accounts and fill empty line_account fields."""
    conn = get_db()
    rows = conn.execute(
        "SELECT o.shopee_account, o.seller_note "
        "FROM orders o "
        "JOIN customers c ON c.shopee_account = o.shopee_account "
        "WHERE (c.line_account IS NULL OR c.line_account = '') "
        "  AND o.seller_note IS NOT NULL AND o.seller_note != ''"
    ).fetchall()
    updates = {}
    for acct, note in rows:
        candidate = _extract_line_candidate(note)
        if candidate and acct not in updates:
            updates[acct] = candidate
    for acct, line in updates.items():
        conn.execute(
            "UPDATE customers SET line_account=? "
            "WHERE shopee_account=? AND (line_account IS NULL OR line_account='')",
            (line, acct))
    conn.commit()
    conn.close()
    return len(updates)


# ── Image OCR via Claude API ──────────────────────────────────────────────────

_SETTINGS_PATH = os.path.join(BASE_DIR, 'data', 'settings.json')


def _load_settings() -> dict:
    if os.path.exists(_SETTINGS_PATH):
        try:
            with open(_SETTINGS_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_settings(s: dict):
    os.makedirs(os.path.dirname(_SETTINGS_PATH), exist_ok=True)
    with open(_SETTINGS_PATH, 'w', encoding='utf-8') as f:
        json.dump(s, f, ensure_ascii=False, indent=2)


def _get_api_auth():
    """Return (api_type, credential) or None.
    api_type: 'gemini' | 'anthropic' | 'anthropic-oauth'
    Priority: Gemini (free) → Anthropic API key → Claude Code OAuth."""
    gemini_key = os.environ.get('GEMINI_API_KEY', '').strip()
    if not gemini_key:
        gemini_key = _load_settings().get('gemini_api_key', '').strip()
    if gemini_key:
        return ('gemini', gemini_key)

    key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
    if key:
        return ('anthropic', key)

    key = _load_settings().get('anthropic_api_key', '').strip()
    if key:
        return ('anthropic', key)

    creds_path = os.path.join(os.path.expanduser('~'), '.claude', '.credentials.json')
    if os.path.exists(creds_path):
        try:
            with open(creds_path, 'r', encoding='utf-8') as f:
                creds = json.load(f)
            token = (creds.get('claudeAiOauth') or {}).get('accessToken', '')
            if token:
                return ('anthropic-oauth', token)
        except Exception:
            pass

    return None


_OCR_PROMPT = (
    '這是一張蝦皮超商取件標籤。版面分上下兩區：\n\n'
    '【上半區】含路線碼（如 X01-2-4）、日期、門市交寄、訂單編號、寄件編號、條碼。'
    '右側小框內是「賣家」姓名，格式為中間有星號遮蔽（例如 林*蕙、林*薇）。\n\n'
    '【下半區】含超商店名、QR code、圓框地區字（南/北/中/東/西等），'
    '圓框右側是「收件人」完整真實姓名（2～4 個中文字，無星號）以及取件金額數字。\n\n'
    '請辨識以下兩項：\n\n'
    '1. 訂單編號：上半區明確標示「訂單編號:」後面的那串英數字（通常以日期數字開頭，'
    '例如 2510130NBW55XV 或 251124JHPEEE6W）。\n'
    '   注意：「寄件編號:」後的 TW 開頭字串是物流單號，不是訂單編號，請勿填入。\n\n'
    '2. 收件人姓名：下半區圓框右側的完整中文姓名。\n'
    '   注意：右上角小框裡含星號的名字（如 林*蕙）是賣家，絕對不可填入此欄。\n\n'
    '只回傳一行 JSON，不含任何說明文字：\n'
    '{"order_id": "...", "real_name": "..."}\n'
    '若無法辨識某欄位，填 null。'
)


def _parse_ocr_json(text: str) -> dict:
    m = re.search(r'\{[^{}]+\}', text)
    if not m:
        raise RuntimeError(f'無法解析 API 回應：{text[:120]}')
    return json.loads(m.group())


def _ocr_with_gemini(b64: str, mime: str, key: str) -> dict:
    import time as _time
    url = f'https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={key}'
    payload = {
        'contents': [{'parts': [
            {'inline_data': {'mime_type': mime, 'data': b64}},
            {'text': _OCR_PROMPT},
        ]}],
        'generationConfig': {'maxOutputTokens': 256, 'temperature': 0},
    }
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode('utf-8'),
        headers={'content-type': 'application/json'},
        method='POST')
    last_err = None
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                body = json.loads(resp.read().decode('utf-8'))
            last_err = None
            break
        except urllib.error.HTTPError as e:
            if e.code in (400, 401, 403):
                raise RuntimeError('AUTH_INVALID')
            if e.code == 429 and attempt < 2:
                _time.sleep(10)
                continue
            detail = e.read().decode('utf-8', errors='replace')[:300]
            last_err = RuntimeError(f'Gemini API 錯誤 {e.code}：{detail}')
        except TimeoutError:
            last_err = RuntimeError('連線逾時（30 秒），請確認網路連線後重試。')
    if last_err:
        raise last_err
    text = ((body.get('candidates') or [{}])[0]
            .get('content', {}).get('parts', [{}])[0].get('text', '')).strip()
    return _parse_ocr_json(text)


def _ocr_with_anthropic(b64: str, mime: str, api_type: str, credential: str) -> dict:
    import time as _time
    if api_type == 'anthropic-oauth':
        headers = {
            'Authorization': f'Bearer {credential}',
            'anthropic-version': '2023-06-01',
            'content-type': 'application/json',
        }
    else:
        headers = {
            'x-api-key': credential,
            'anthropic-version': '2023-06-01',
            'content-type': 'application/json',
        }
    payload = {
        'model': 'claude-haiku-4-5-20251001',
        'max_tokens': 256,
        'messages': [{'role': 'user', 'content': [
            {'type': 'image', 'source': {'type': 'base64', 'media_type': mime, 'data': b64}},
            {'type': 'text', 'text': _OCR_PROMPT},
        ]}],
    }
    req = urllib.request.Request(
        'https://api.anthropic.com/v1/messages',
        data=json.dumps(payload).encode('utf-8'),
        headers=headers,
        method='POST')
    last_err = None
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                body = json.loads(resp.read().decode('utf-8'))
            last_err = None
            break
        except urllib.error.HTTPError as e:
            if e.code in (401, 403):
                raise RuntimeError('AUTH_INVALID')
            if e.code == 429 and attempt < 2:
                _time.sleep(8)
                continue
            detail = e.read().decode('utf-8', errors='replace')[:300]
            last_err = RuntimeError(f'API 回傳錯誤 {e.code}：{detail}')
        except TimeoutError:
            last_err = RuntimeError('連線逾時（30 秒），請確認網路連線後重試。')
    if last_err:
        raise last_err
    text = (body.get('content') or [{}])[0].get('text', '').strip()
    return _parse_ocr_json(text)


def _ocr_shipping_label(image_path: str) -> dict:
    """Extract order_id + real_name from a shipping label image.
    Returns {'order_id': str|None, 'real_name': str|None}.
    Raises RuntimeError('AUTH_MISSING') | RuntimeError('AUTH_INVALID') | RuntimeError(msg)."""
    auth = _get_api_auth()
    if not auth:
        raise RuntimeError('AUTH_MISSING')
    api_type, credential = auth

    with open(image_path, 'rb') as f:
        raw = f.read()
    b64 = base64.b64encode(raw).decode('ascii')
    ext = os.path.splitext(image_path)[1].lower()
    mime = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
            '.png': 'image/png', '.webp': 'image/webp'}.get(ext, 'image/jpeg')

    if api_type == 'gemini':
        return _ocr_with_gemini(b64, mime, credential)
    else:
        return _ocr_with_anthropic(b64, mime, api_type, credential)


# ── Auto-update ───────────────────────────────────────────────────────────────

def _ver_tuple(v: str):
    try:    return tuple(int(x) for x in str(v).strip().split('.'))
    except: return (0,)


def _fetch_update_info() -> dict | None:
    """Fetch update.json from UPDATE_CHECK_URL. Returns dict or None on any failure."""
    url = UPDATE_CHECK_URL.strip()
    if not url:
        return None
    try:
        req = urllib.request.Request(url, headers={'User-Agent': f'JadeSystem/{APP_VERSION}'})
        with urllib.request.urlopen(req, timeout=8) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except Exception:
        return None


def _do_update(root, app_url: str, new_ver: str):
    """Download new app.py and replace current file. Shows result dialog."""
    this_file  = os.path.abspath(__file__)
    tmp_path   = this_file + '.new'
    backup     = this_file + '.bak'
    try:
        req = urllib.request.Request(app_url, headers={'User-Agent': f'JadeSystem/{APP_VERSION}'})
        with urllib.request.urlopen(req, timeout=60) as resp:
            new_code = resp.read()

        # Basic sanity check — must look like a Python file
        snippet = new_code[:500].decode('utf-8', errors='replace')
        if 'import' not in snippet and 'def ' not in snippet:
            raise ValueError('下載的檔案內容異常，請稍後再試。')

        with open(tmp_path, 'wb') as f:
            f.write(new_code)

        # Atomic-ish replace: backup current → put new in place
        if os.path.exists(backup):
            os.remove(backup)
        shutil.copy2(this_file, backup)
        shutil.move(tmp_path, this_file)

        ans = messagebox.askyesno(
            '更新完成',
            f'已更新至版本 {new_ver}。\n\n'
            f'請關閉程式後重新開啟以套用更新。\n\n'
            f'是否立即關閉？',
            parent=root)
        if ans:
            root.destroy()

    except Exception as e:
        if os.path.exists(tmp_path):
            try: os.remove(tmp_path)
            except Exception: pass
        messagebox.showerror('更新失敗', f'下載更新時發生錯誤：\n{e}', parent=root)


def _on_update_available(root, info: dict):
    """Called on main thread when a newer version is found."""
    new_ver  = (info.get('version') or '').strip()
    app_url  = (info.get('app_url') or '').strip()
    changes  = info.get('changelog') or []
    if isinstance(changes, str):
        changes = [changes]

    if _ver_tuple(new_ver) <= _ver_tuple(APP_VERSION):
        return

    bullet = '\n'.join(f'  • {c}' for c in changes) or '  （無詳細說明）'
    ans = messagebox.askyesno(
        f'發現新版本 {new_ver}',
        f'目前版本：{APP_VERSION}　→　新版本：{new_ver}\n\n'
        f'此次更新內容：\n{bullet}\n\n'
        f'是否立即下載並更新？',
        parent=root)
    if ans:
        if app_url:
            _do_update(root, app_url, new_ver)
        else:
            messagebox.showwarning('無法更新', '未設定下載連結，請聯絡系統管理員。', parent=root)


def _on_announcement(root, text: str, seen_key: str):
    """Show announcement once per session (tracked by seen_key)."""
    settings = _load_settings()
    if settings.get('_last_announcement') == seen_key:
        return
    settings['_last_announcement'] = seen_key
    _save_settings(settings)
    messagebox.showinfo('系統公告', text, parent=root)


def _start_update_check(root):
    """Launch background thread to check for updates/announcements."""
    def worker():
        info = _fetch_update_info()
        if not info:
            return
        announcement = (info.get('announcement') or '').strip()
        new_ver      = (info.get('version')      or '').strip()

        delay = 1200  # ms — let the UI fully load first
        if announcement:
            key = f'{new_ver}|{hash(announcement)}'
            root.after(delay, lambda: _on_announcement(root, announcement, key))
            delay += 300
        if new_ver and _ver_tuple(new_ver) > _ver_tuple(APP_VERSION):
            root.after(delay, lambda: _on_update_available(root, info))

    threading.Thread(target=worker, daemon=True).start()


# ── Finance helpers ───────────────────────────────────────────────────────────

def _get_refunded(o):
    try:    return o['refunded_price'] or 0.0
    except: return 0.0


def _calc_net(o):
    refunded = _get_refunded(o)
    # Legacy data (old imports): refunded_price=0 but refund_status set → fully refunded
    if refunded == 0 and (o['refund_status'] or '').strip():
        return -(o['return_shipping'] or 0.0)
    effective = (o['original_price'] or 0.0) - refunded
    return (effective
            - (o['seller_coupon']   or 0.0)
            - (o['transaction_fee'] or 0.0)
            - (o['service_fee']     or 0.0)
            - (o['payment_fee']     or 0.0)
            - (o['return_shipping'] or 0.0))


_NET_SQL = (
    "CASE "
    # Legacy: refund_status set but refunded_price not yet tracked → old full-refund behaviour
    "  WHEN COALESCE(refunded_price,0)=0 AND refund_status!='' AND refund_status IS NOT NULL "
    "    THEN -return_shipping "
    # New data: effective = original - refunded; handles clean / partial / full refund uniformly
    "  ELSE (original_price - COALESCE(refunded_price,0)) "
    "       - seller_coupon - transaction_fee - service_fee - payment_fee - return_shipping "
    "END"
)


# ── Tooltip ───────────────────────────────────────────────────────────────────

class _Tooltip:
    def __init__(self, widget, get_text_fn):
        self._w   = widget
        self._fn  = get_text_fn
        self._tip = None
        self._lbl = None
        widget.bind('<Motion>', self._show)
        widget.bind('<Leave>',  self._hide)
        widget.bind('<Button>', self._hide)

    def _show(self, e):
        text = self._fn(e)
        if not text:
            self._hide()
            return
        x, y = e.x_root + 14, e.y_root + 14
        if self._tip and self._tip.winfo_exists():
            self._tip.geometry(f'+{x}+{y}')
            self._lbl.config(text=text)
        else:
            self._tip = tk.Toplevel(self._w)
            self._tip.wm_overrideredirect(True)
            self._tip.wm_geometry(f'+{x}+{y}')
            self._lbl = tk.Label(
                self._tip, text=text, bg='#FFFDE7', fg='#333',
                font=(FONT, 9), relief='solid', bd=1,
                padx=8, pady=4, wraplength=460, justify='left')
            self._lbl.pack()

    def _hide(self, *_):
        if self._tip and self._tip.winfo_exists():
            self._tip.destroy()
        self._tip = None
        self._lbl = None


# ── Chart Window ──────────────────────────────────────────────────────────────

class _ChartWindow(tk.Toplevel):
    """折線統計圖：銷售總額 + 進帳總額，依年/月/日粒度顯示。"""

    def __init__(self, master, start_date=None, end_date=None):
        super().__init__(master)
        self.title('銷售統計折線圖')
        self.geometry('900x520')
        self.minsize(600, 360)
        self.configure(bg=WHITE)
        self.transient(master)
        self._start = start_date
        self._end   = end_date
        self._gran  = tk.StringVar(value='月')
        self._chart_xs   = []
        self._chart_data = []
        self._build()
        self.after(80, self._refresh)

    def _build(self):
        top = tk.Frame(self, bg=WHITE, pady=10)
        top.pack(fill='x', padx=16)

        tk.Label(top, text='銷售統計折線圖', bg=WHITE, fg=JADE,
                 font=(FONT, 13, 'bold')).pack(side='left')

        right = tk.Frame(top, bg=WHITE)
        right.pack(side='right')
        tk.Label(right, text='時間粒度：', bg=WHITE, fg=TEXT, font=(FONT, 10)).pack(side='left')
        for label in ('年', '月', '日'):
            rb = tk.Radiobutton(right, text=label, variable=self._gran, value=label,
                                command=self._refresh, bg=WHITE, fg=TEXT,
                                activebackground=JADE_MID, selectcolor=JADE_MID,
                                font=(FONT, 10))
            rb.pack(side='left', padx=4)

        # Date range display
        range_txt = ''
        if self._start or self._end:
            range_txt = f'  篩選：{self._start or "─"} ～ {self._end or "─"}'
        if range_txt:
            tk.Label(right, text=range_txt, bg=WHITE, fg=GRAY,
                     font=(FONT, 9)).pack(side='left', padx=(12, 0))

        self._canvas = tk.Canvas(self, bg='#FAFFFE', highlightthickness=0)
        self._canvas.pack(fill='both', expand=True, padx=12, pady=(0, 12))
        self._canvas.bind('<Configure>', lambda _: self._refresh())
        self._canvas.bind('<Motion>',   self._on_hover)
        self._canvas.bind('<Leave>',    self._on_leave)

    def _refresh(self):
        gran   = self._gran.get()
        fmt    = {'年': '%Y', '月': '%Y-%m', '日': '%Y-%m-%d'}[gran]
        where  = "WHERE order_status != '不成立'"
        params = []
        if self._start:
            where += ' AND order_date >= ?'; params.append(self._start)
        if self._end:
            where += ' AND order_date <= ?'; params.append(self._end + ' 23:59:59')

        conn = get_db()
        rows = conn.execute(f'''
            SELECT strftime('{fmt}', order_date) period,
                   COALESCE(SUM(CASE WHEN refund_status='' OR refund_status IS NULL
                                THEN original_price ELSE 0 END), 0) sales,
                   COALESCE(SUM({_NET_SQL}), 0) income
            FROM orders {where}
            GROUP BY period
            ORDER BY period
        ''', params).fetchall()
        conn.close()

        self._chart_data = [{'period': r[0], 'sales': r[1], 'income': r[2]}
                            for r in rows if r[0]]
        self._draw()

    def _draw(self):
        c = self._canvas
        c.delete('all')
        W = c.winfo_width()
        H = c.winfo_height()
        if W < 80 or H < 80:
            return

        data = self._chart_data
        if not data:
            c.create_text(W // 2, H // 2, text='此時段無資料',
                          font=(FONT, 14), fill=GRAY)
            return

        ML, MR, MT, MB = 100, 140, 48, 56

        periods = [d['period'] for d in data]
        sales   = [d['sales']  for d in data]
        income  = [d['income'] for d in data]
        n = len(periods)

        max_v = max(max(sales, default=0), max(income, default=0)) or 1
        min_v = min(min(income, default=0), 0)
        rng   = max_v - min_v or 1
        max_v += rng * 0.12
        if min_v < 0:
            min_v -= rng * 0.06
        rng = max_v - min_v

        iw = (W - ML - MR) / max(n - 1, 1) if n > 1 else 0
        xs = [ML + i * iw for i in range(n)]
        self._chart_xs = xs

        def ty(v):
            return MT + (max_v - v) / rng * (H - MT - MB)

        # ── Grid & Y axis labels ─────────────────────────────────────────
        for i in range(5):
            frac = i / 4
            y = MT + frac * (H - MT - MB)
            v = max_v - frac * rng
            c.create_line(ML, y, W - MR, y, fill='#EBEBEB', dash=(4, 3))
            c.create_text(ML - 8, y, text=f'NT${v:,.0f}',
                          anchor='e', font=(FONT, 8), fill='#999')

        # ── Axes ─────────────────────────────────────────────────────────
        c.create_line(ML, MT, ML, H - MB, fill='#CCCCCC', width=1)
        c.create_line(ML, H - MB, W - MR, H - MB, fill='#CCCCCC', width=1)

        # ── X axis labels ─────────────────────────────────────────────────
        step = max(1, n // 14)
        for i, (x, p) in enumerate(zip(xs, periods)):
            if i == 0 or i == n - 1 or i % step == 0:
                c.create_text(x, H - MB + 14, text=p,
                              font=(FONT, 8), fill='#999', anchor='n')

        # ── Draw line helper ──────────────────────────────────────────────
        def draw_series(values, color, tag):
            pts = [(xs[i], ty(values[i])) for i in range(n)]
            if n >= 2:
                flat = [v for pt in pts for v in pt]
                c.create_line(*flat, fill=color, width=2,
                              smooth=(n >= 5), tags=tag)
            for px, py in pts:
                c.create_oval(px - 3, py - 3, px + 3, py + 3,
                              fill=color, outline=WHITE, width=1, tags=tag)

        draw_series(income, JADE, 'income')
        draw_series(sales,  BLUE, 'sales')

        # ── Legend ───────────────────────────────────────────────────────
        lx = W - MR + 10
        c.create_rectangle(lx, MT, W - 6, MT + 58,
                           fill=WHITE, outline='#DDDDDD')
        for idx2, (col, lbl) in enumerate([(BLUE, '銷售總額'), (JADE, '進帳總額')]):
            ly = MT + 16 + idx2 * 24
            c.create_line(lx + 8,  ly, lx + 28, ly, fill=col, width=2)
            c.create_oval(lx + 15, ly - 3, lx + 21, ly + 3,
                          fill=col, outline=WHITE)
            c.create_text(lx + 34, ly, text=lbl, anchor='w',
                          font=(FONT, 9), fill=TEXT)

    def _on_hover(self, e):
        if not self._chart_xs or not self._chart_data:
            return
        # Find nearest point by X
        dists = [abs(x - e.x) for x in self._chart_xs]
        idx   = dists.index(min(dists))
        x     = self._chart_xs[idx]
        d     = self._chart_data[idx]

        c = self._canvas
        c.delete('hover_overlay')
        W = c.winfo_width()
        H = c.winfo_height()
        MB = 56

        c.create_line(x, 44, x, H - MB, fill='#BBBBBB',
                      dash=(3, 3), tags='hover_overlay')

        tip_w, tip_h = 162, 62
        tx = x + 10 if x < W - 185 else x - tip_w - 10
        ty_box = 50
        c.create_rectangle(tx, ty_box, tx + tip_w, ty_box + tip_h,
                           fill='#FFFDE7', outline='#CCCCCC', tags='hover_overlay')
        c.create_text(tx + 8, ty_box + 8, text=d['period'], anchor='nw',
                      font=(FONT, 9, 'bold'), fill=TEXT, tags='hover_overlay')
        c.create_text(tx + 8, ty_box + 24,
                      text=f'銷售：NT${d["sales"]:,.0f}', anchor='nw',
                      font=(FONT, 9), fill=BLUE, tags='hover_overlay')
        c.create_text(tx + 8, ty_box + 40,
                      text=f'進帳：NT${d["income"]:,.0f}', anchor='nw',
                      font=(FONT, 9), fill=JADE, tags='hover_overlay')

    def _on_leave(self, _e):
        self._canvas.delete('hover_overlay')


# ── Date Range Picker ─────────────────────────────────────────────────────────

class _DateRangePicker(tk.Toplevel):
    """日期範圍選擇器：左右雙日曆，選好後套用篩選。"""

    _CAL_KW = dict(
        selectmode='day',
        date_pattern='yyyy-mm-dd',
        showweeknumbers=False,
        firstweekday='sunday',
        font=(FONT, 9),
        background=JADE_DARK,
        foreground=WHITE,
        headersbackground=JADE_MID,
        headersforeground=JADE,
        selectbackground=JADE,
        selectforeground=WHITE,
        normalforeground=TEXT,
        weekendforeground=RED,
        othermonthforeground=GRAY,
    )

    def __init__(self, master, start_var, end_var):
        super().__init__(master)
        self.title('選擇日期範圍')
        self.configure(bg=WHITE)
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self._sv = start_var
        self._ev = end_var
        self._build()
        # Center over parent
        self.update_idletasks()
        px = master.winfo_x() + (master.winfo_width()  - self.winfo_width())  // 2
        py = master.winfo_y() + (master.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{max(px,0)}+{max(py,0)}')

    def _build(self):
        import datetime
        today = datetime.date.today()

        tk.Label(self, text='選擇日期範圍', bg=WHITE, fg=JADE,
                 font=(FONT, 13, 'bold')).pack(pady=(14, 2))
        tk.Label(self, text='左側選開始日期，右側選結束日期，再點「套用」',
                 bg=WHITE, fg=GRAY, font=(FONT, 9)).pack(pady=(0, 10))

        cals_frame = tk.Frame(self, bg=WHITE)
        cals_frame.pack(padx=20, pady=(0, 6))

        def _parse(s):
            try:
                return datetime.date.fromisoformat(s) if s else today
            except ValueError:
                return today

        d_start = _parse(self._sv.get().strip())
        d_end   = _parse(self._ev.get().strip())

        for col_side, label, attr, init_d in (
            ('left',  '開始日期', '_cal_s', d_start),
            ('right', '結束日期', '_cal_e', d_end),
        ):
            col = tk.Frame(cals_frame, bg=WHITE)
            col.pack(side=col_side, padx=12)
            hdr = tk.Frame(col, bg=JADE_MID)
            hdr.pack(fill='x', pady=(0, 4))
            tk.Label(hdr, text=label, bg=JADE_MID, fg=JADE,
                     font=(FONT, 10, 'bold'), pady=5).pack()
            cal = Calendar(col, year=init_d.year, month=init_d.month, day=init_d.day,
                           **_DateRangePicker._CAL_KW)
            cal.pack()
            # Restore previously selected date
            try:
                cal.selection_set(init_d.strftime('%Y-%m-%d'))
            except Exception:
                pass
            setattr(self, attr, cal)

        # Live preview label
        self._preview = tk.Label(self, text='', bg=WHITE, fg=JADE,
                                 font=(FONT, 10, 'bold'))
        self._preview.pack(pady=(4, 0))

        def _update_preview(*_):
            s = self._cal_s.get_date()
            e = self._cal_e.get_date()
            ok = s <= e
            color = JADE if ok else RED
            hint  = '' if ok else '  ⚠ 結束日期早於開始日期'
            self._preview.config(text=f'已選：{s} ～ {e}{hint}', fg=color)

        self._cal_s.bind('<<CalendarSelected>>', _update_preview)
        self._cal_e.bind('<<CalendarSelected>>', _update_preview)
        _update_preview()

        # Buttons
        bf = tk.Frame(self, bg=WHITE)
        bf.pack(pady=(8, 16))

        def confirm():
            s = self._cal_s.get_date()
            e = self._cal_e.get_date()
            if s > e:
                messagebox.showwarning('日期錯誤', '開始日期不能晚於結束日期',
                                       parent=self)
                return
            self._sv.set(s)
            self._ev.set(e)
            self.destroy()

        def clear():
            self._sv.set('')
            self._ev.set('')
            self.destroy()

        App._btn(bf, '✓ 套用篩選', JADE, confirm).pack(side='left', padx=8)
        App._btn(bf, '清除篩選',   RED,  clear  ).pack(side='left', padx=8)
        App._btn(bf, '取消',       GRAY, self.destroy).pack(side='left', padx=8)


# ── GUI ───────────────────────────────────────────────────────────────────────

_HDR_TEXTS = {
    'acct':  '蝦皮帳號',
    'name':  '姓名',
    'line':  'LINE',
    'cnt':   '訂單',
    'spent': '消費金額',
    'last':  '最近訂購',
}


class App(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title('瑾琰上品')
        self.geometry('1300x760')
        self.minsize(950, 620)
        self.configure(bg=BG)
        self._current    = None
        self._field_vars = {}
        self._pref_text  = None
        self._notes_text = None
        self._sort_col   = None
        self._sort_dir   = 'asc'
        # Date filter vars (must exist before _build_toolbar)
        self.start_date_var = tk.StringVar()
        self.end_date_var   = tk.StringVar()
        self.start_date_var.trace_add('write', lambda *_: self._refresh_list())
        self.end_date_var.trace_add('write',   lambda *_: self._refresh_list())
        self._setup_styles()
        self._build_ui()
        # Make stats labels clickable for chart
        for attr in ('lbl_sales', 'lbl_rev'):
            lbl = getattr(self, attr)
            lbl.config(cursor='hand2')
            lbl.bind('<Button-1>', lambda e: self._open_chart())
            lbl.bind('<Enter>',    lambda e, l=lbl: l.config(fg='#FFFFFF'))
            lbl.bind('<Leave>',    lambda e, l=lbl: l.config(fg='#B7E4C7'))
        self._refresh_list()
        _scan_line_from_notes()
        _start_update_check(self)

    # ── Style ──────────────────────────────────────────────────────────────

    def _setup_styles(self):
        s = ttk.Style(self)
        s.theme_use('clam')
        s.configure('TFrame',    background=BG)
        s.configure('TLabel',    background=BG, foreground=TEXT, font=(FONT, 10))
        s.configure('TButton',   font=(FONT, 10))
        s.configure('TCombobox', font=(FONT, 10))
        s.configure('Treeview',
                    background=WHITE, fieldbackground=WHITE,
                    foreground=TEXT,  font=(FONT, 10), rowheight=28)
        s.configure('Treeview.Heading',
                    background=JADE_MID, foreground=JADE,
                    font=(FONT, 10, 'bold'), relief='flat')
        s.map('Treeview',
              background=[('selected', JADE)],
              foreground=[('selected', WHITE)])
        s.map('Treeview.Heading', background=[('active', JADE_MID)])
        s.configure('TSeparator', background='#DDD')

    # ── Build UI ────────────────────────────────────────────────────────────

    def _build_ui(self):
        self._build_header()
        self._build_toolbar()
        self._build_pane()
        self._build_statusbar()

    def _build_header(self):
        hdr = tk.Frame(self, bg=JADE, height=54)
        hdr.pack(fill='x')
        hdr.pack_propagate(False)
        tk.Label(hdr, text='  瑾琰上品', bg=JADE, fg=WHITE,
                 font=(FONT, 16, 'bold')).pack(side='left', padx=4)
        tk.Label(hdr, text=f'v{APP_VERSION}', bg=JADE, fg='#8EC4A8',
                 font=(FONT, 8)).pack(side='left', padx=(0, 8), pady=(14, 0))
        # Right-to-left: rev, sales, refund, all, cust
        for attr in ('lbl_rev', 'lbl_sales', 'lbl_refund', 'lbl_all', 'lbl_cust'):
            lbl = tk.Label(hdr, text='', bg=JADE, fg='#B7E4C7', font=(FONT, 10))
            lbl.pack(side='right', padx=12)
            setattr(self, attr, lbl)

    def _build_toolbar(self):
        bar = tk.Frame(self, bg='#EAF4EE', pady=7)
        bar.pack(fill='x')

        # ── Left: search + sort ──────────────────────────────────────────
        tk.Label(bar, text='搜尋：', bg='#EAF4EE', font=(FONT, 10)).pack(side='left', padx=(14, 2))
        self.search_var = tk.StringVar()
        self.search_var.trace_add('write', lambda *_: self._refresh_list())
        tk.Entry(bar, textvariable=self.search_var, font=(FONT, 10), width=18,
                 relief='solid', bd=1).pack(side='left', padx=4)

        tk.Label(bar, text='排序：', bg='#EAF4EE', font=(FONT, 10)).pack(side='left', padx=(8, 2))
        self.sort_var = tk.StringVar(value='最近訂購')
        cb = ttk.Combobox(bar, textvariable=self.sort_var, state='readonly', width=8,
                          values=['最近訂購', '消費金額', '訂單數量', '姓名'])
        cb.pack(side='left', padx=4)
        cb.bind('<<ComboboxSelected>>', self._on_sort_combo)

        # ── Left: date filter (calendar picker) ─────────────────────────
        date_area = tk.Frame(bar, bg='#EAF4EE')
        date_area.pack(side='left', padx=(12, 0))
        self._btn(date_area, '日曆篩選', JADE_DARK, self._open_date_picker,
                  small=True).pack(side='left', padx=(0, 6))
        self._filter_lbl = tk.Label(date_area, text='', bg='#EAF4EE', fg=RED,
                                    font=(FONT, 9))
        self._filter_lbl.pack(side='left')
        self._clear_date_btn = self._btn(date_area, '✕ 清除', '#CC3333',
                                         self._clear_date_filter, small=True)
        # Shown/hidden dynamically by _refresh_list

        # ── Right: action buttons ────────────────────────────────────────
        self._btn(bar, '＋ 新增客戶', JADE, self._dlg_new_customer).pack(side='right', padx=(4, 14))
        self._btn(bar, '↑ 匯入報表',  BLUE, self._import_excel).pack(side='right', padx=4)
        self._btn(bar, '↑ 匯入圖片', '#7B3F9E', self._import_images).pack(side='right', padx=4)
        self._btn(bar, '⚙ API 設定',  '#888888', self._show_api_key_dialog,
                  small=True).pack(side='right', padx=(0, 2))

    def _build_pane(self):
        pw = ttk.PanedWindow(self, orient='horizontal')
        pw.pack(fill='both', expand=True, padx=10, pady=6)
        left  = ttk.Frame(pw)
        right = ttk.Frame(pw)
        pw.add(left,  weight=38)
        pw.add(right, weight=62)
        self._build_list(left)
        self._build_detail(right)

    def _build_list(self, parent):
        cols = ('acct', 'name', 'line', 'cnt', 'spent', 'last')
        self.tree = ttk.Treeview(parent, columns=cols, show='headings', selectmode='browse')
        hdrs = [('acct','蝦皮帳號',130), ('name','姓名',95), ('line','LINE',95),
                ('cnt','訂單',52), ('spent','消費金額',105), ('last','最近訂購',95)]
        for cid, text, w in hdrs:
            self.tree.heading(cid, text=text, command=lambda c=cid: self._sort_by(c))
            anchor = 'e' if cid == 'spent' else ('center' if cid == 'cnt' else 'w')
            self.tree.column(cid, width=w, minwidth=40, anchor=anchor)
        vsb = ttk.Scrollbar(parent, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)
        self.tree.bind('<<TreeviewSelect>>', self._on_select)
        self.tree.bind('<Button-3>', self._on_list_rightclick)

    def _build_detail(self, parent):
        self._cv = tk.Canvas(parent, bg=BG, highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient='vertical', command=self._cv.yview)
        self._cv.configure(yscrollcommand=vsb.set)
        vsb.pack(side='right', fill='y')
        self._cv.pack(side='left', fill='both', expand=True)
        self._df = tk.Frame(self._cv, bg=BG)
        self._df_id = self._cv.create_window((0, 0), window=self._df, anchor='nw')
        self._df.bind('<Configure>', lambda _: (
            self._cv.configure(scrollregion=self._cv.bbox('all')),
        ))
        self._cv.bind('<Configure>', lambda e:
            self._cv.itemconfigure(self._df_id, width=e.width))
        def _mwheel(e): self._cv.yview_scroll(-1*(e.delta//120), 'units')
        self._cv.bind('<Enter>', lambda _: self._cv.bind_all('<MouseWheel>', _mwheel))
        self._cv.bind('<Leave>', lambda _: self._cv.unbind_all('<MouseWheel>'))
        self._show_placeholder()

    def _build_statusbar(self):
        sb = tk.Frame(self, bg='#DDEEE5', height=24)
        sb.pack(fill='x', side='bottom')
        sb.pack_propagate(False)
        self.status_var = tk.StringVar(value='就緒')
        tk.Label(sb, textvariable=self.status_var, bg='#DDEEE5',
                 fg=GRAY, font=(FONT, 9), anchor='w').pack(side='left', padx=10, fill='y')
        self._undo_btn = tk.Button(
            sb, text='', bg='#E74C3C', fg=WHITE,
            font=(FONT, 9, 'bold'), relief='flat', cursor='hand2',
            padx=8, pady=0, command=self._undo_delete,
            activebackground='#C0392B', activeforeground=WHITE)
        self._undo_backup  = None
        self._undo_job     = None
        self._undo_seconds = 0

    # ── Sorting ────────────────────────────────────────────────────────────

    def _sort_by(self, col):
        if self._sort_col == col:
            self._sort_dir = 'desc' if self._sort_dir == 'asc' else 'asc'
        else:
            self._sort_col = col
            self._sort_dir = 'asc'
        self._refresh_list()

    def _on_sort_combo(self, _event):
        self._sort_col = None
        self._refresh_list()

    def _update_heading_arrows(self):
        for cid, base in _HDR_TEXTS.items():
            arrow = (' ▲' if self._sort_dir == 'asc' else ' ▼') if cid == self._sort_col else ''
            self.tree.heading(cid, text=base + arrow)

    # ── Detail Panel ────────────────────────────────────────────────────────

    def _clear_detail(self):
        for w in self._df.winfo_children():
            w.destroy()
        self._field_vars.clear()
        self._pref_text  = None
        self._notes_text = None

    def _show_placeholder(self):
        self._clear_detail()
        tk.Label(self._df, text='← 從左側選擇客戶以查看詳情',
                 bg=BG, fg='#AAA', font=(FONT, 12)).pack(pady=60)

    @staticmethod
    def _short_status(status):
        if not status:
            return ''
        for kw in ('已完成', '已出貨', '不成立', '退款', '退貨', '處理中', '備貨中', '待出貨'):
            if status.startswith(kw):
                return kw
        return status[:10]

    @staticmethod
    def _product_display(raw_name):
        """Return (cell_text, tooltip_text) for order treeview product column."""
        parts = [s.strip() for s in (raw_name or '').split('|') if s.strip()]
        if not parts:
            return '', ''
        if len(parts) == 1:
            return parts[0][:55], (parts[0] if len(parts[0]) > 20 else '')
        cell = parts[0][:32] + f'  (+{len(parts)-1}件)'
        tip  = '\n'.join(f'{i+1}. {p}' for i, p in enumerate(parts))
        return cell, tip

    def _show_detail(self, acct):
        conn = get_db()
        c   = conn.execute('SELECT * FROM customers WHERE shopee_account=?', (acct,)).fetchone()
        os_ = conn.execute(
            'SELECT * FROM orders WHERE shopee_account=? ORDER BY order_date DESC',
            (acct,)).fetchall()
        st  = conn.execute(
            f'SELECT COUNT(*) t, '
            f'COUNT(CASE WHEN order_status!="不成立" AND '
            f'      (refund_status="" OR refund_status IS NULL '
            f'       OR (COALESCE(refunded_price,0)>0 AND refunded_price<original_price)) THEN 1 END) ok, '
            f'COALESCE(SUM(CASE WHEN order_status!="不成立" THEN {_NET_SQL} ELSE 0 END),0) net_rev, '
            f'COALESCE(SUM(CASE WHEN order_status!="不成立" '
            f'               AND (COALESCE(refunded_price,0)>0 OR refund_status="" OR refund_status IS NULL) '
            f'               THEN (original_price - COALESCE(refunded_price,0)) ELSE 0 END),0) sales_rev '
            f'FROM orders WHERE shopee_account=?', (acct,)).fetchone()
        conn.close()
        if not c:
            return

        self._clear_detail()

        # ── Info card ─────────────────────────────────────────────────────
        card = tk.Frame(self._df, bg=WHITE, bd=0)
        card.pack(fill='x', padx=12, pady=(12, 4))

        name_display = c['real_name'] or '（未設定姓名）'
        vip = '  ★ VIP' if (st['ok'] >= 5 and st['sales_rev'] >= 10000) else ''
        tk.Label(card, text=f'{name_display}{vip}', bg=WHITE, fg=JADE,
                 font=(FONT, 14, 'bold')).pack(anchor='w', padx=16, pady=(12, 0))

        acct_row = tk.Frame(card, bg=WHITE)
        acct_row.pack(anchor='w', padx=16, pady=(0, 8))
        tk.Label(acct_row, text='蝦皮帳號：', bg=WHITE, fg=GRAY, font=(FONT, 9)).pack(side='left')
        _ae = tk.Entry(acct_row, font=(FONT, 9), relief='flat', fg=GRAY,
                       readonlybackground=WHITE, bd=0, width=len(acct) + 2)
        _ae.insert(0, acct)
        _ae.config(state='readonly')
        _ae.pack(side='left')

        strip = tk.Frame(card, bg=JADE_MID)
        strip.pack(fill='x', padx=16, pady=(0, 10), ipady=7)
        tk.Label(strip, text=f'完成訂單  {st["ok"]}', bg=JADE_MID, fg=JADE,
                 font=(FONT, 10, 'bold')).pack(side='left', padx=16)
        tk.Label(strip, text=f'銷售總額  NT$ {st["sales_rev"]:,.0f}', bg=JADE_MID, fg=BLUE,
                 font=(FONT, 10, 'bold')).pack(side='left', padx=16)
        tk.Label(strip, text=f'進帳總額  NT$ {st["net_rev"]:,.0f}', bg=JADE_MID, fg=JADE_DARK,
                 font=(FONT, 10, 'bold')).pack(side='left', padx=16)
        tk.Label(strip, text=f'全部訂單  {st["t"]}', bg=JADE_MID, fg=GRAY,
                 font=(FONT, 10)).pack(side='left', padx=16)

        ttk.Separator(card, orient='horizontal').pack(fill='x', padx=16, pady=(0, 8))

        form = tk.Frame(card, bg=WHITE)
        form.pack(fill='x', padx=16, pady=(0, 4))
        form.columnconfigure(1, weight=1)

        for row_i, (lbl, key, val) in enumerate([
            ('真實姓名',  'real_name',    c['real_name']    or ''),
            ('LINE 帳號', 'line_account', c['line_account'] or ''),
            ('生　　日',  'birthday',     c['birthday']     or ''),
        ]):
            tk.Label(form, text=lbl, bg=WHITE, fg=TEXT,
                     font=(FONT, 10), width=9, anchor='e').grid(
                row=row_i, column=0, sticky='e', padx=(0, 8), pady=4)
            var = tk.StringVar(value=val)
            self._field_vars[key] = var
            tk.Entry(form, textvariable=var, font=(FONT, 10),
                     relief='solid', bd=1).grid(row=row_i, column=1, sticky='ew', pady=4)

        tk.Label(form, text='生日格式: YYYY-MM-DD', bg=WHITE, fg='#BBB',
                 font=(FONT, 8)).grid(row=3, column=1, sticky='w')

        tk.Label(form, text='喜好標籤', bg=WHITE, fg=TEXT,
                 font=(FONT, 10), width=9, anchor='e').grid(
            row=4, column=0, sticky='ne', padx=(0, 8), pady=4)
        self._pref_text = tk.Text(form, height=3, font=(FONT, 10),
                                  relief='solid', bd=1, wrap='word')
        self._pref_text.insert('1.0', c['preferences'] or '')
        self._pref_text.grid(row=4, column=1, sticky='ew', pady=4)

        tk.Label(form, text='備　　註', bg=WHITE, fg=TEXT,
                 font=(FONT, 10), width=9, anchor='e').grid(
            row=5, column=0, sticky='ne', padx=(0, 8), pady=4)
        self._notes_text = tk.Text(form, height=3, font=(FONT, 10),
                                   relief='solid', bd=1, wrap='word')
        self._notes_text.insert('1.0', c['notes'] or '')
        self._notes_text.grid(row=5, column=1, sticky='ew', pady=4)

        btn_row = tk.Frame(card, bg=WHITE)
        btn_row.pack(anchor='w', padx=16, pady=(4, 14))
        self._btn(btn_row, '✓ 儲存資料', JADE, self._save).pack(side='left')
        self._btn(btn_row, '✕ 刪除客戶', RED, self._delete_customer).pack(side='left', padx=(12, 0))

        # ── Order history ─────────────────────────────────────────────────
        ohdr_frame = tk.Frame(self._df, bg=WHITE, bd=0)
        ohdr_frame.pack(fill='x', padx=12, pady=(4, 0))
        tk.Label(ohdr_frame,
                 text=f'購買記錄（共 {len(os_)} 筆）｜雙擊列可查看完整內容',
                 bg=WHITE, fg=JADE,
                 font=(FONT, 11, 'bold')).pack(anchor='w', padx=16, pady=(10, 6))

        oc = tk.Frame(self._df, bg=WHITE)
        oc.pack(fill='both', padx=12, pady=(0, 12))

        ocols = ('oid', 'status', 'refund', 'order_date', 'complete_date',
                 'product', 'orig_price', 'net_income',
                 'coupon', 'payment', 'shipping', 'tracking',
                 'city', 'district', 'note')
        ot = ttk.Treeview(oc, columns=ocols, show='headings',
                          height=min(max(len(os_), 4), 12))
        ohdrs = [
            ('oid',          '訂單編號',     145),
            ('status',       '狀態',          72),
            ('refund',       '退貨退款',       80),
            ('order_date',   '訂單成立日期', 110),
            ('complete_date','訂單完成時間', 110),
            ('product',      '商品名稱',     210),
            ('orig_price',   '商品原價',      95),
            ('net_income',   '進帳金額',      95),
            ('coupon',       '優惠券',         72),
            ('payment',      '付款方式',       90),
            ('shipping',     '寄送方式',      110),
            ('tracking',     '包裹查詢號碼',  130),
            ('city',         '城市',           72),
            ('district',     '行政區',         72),
            ('note',         '備註',          120),
        ]
        for cid, text, w in ohdrs:
            ot.heading(cid, text=text)
            anchor = 'e' if cid in ('orig_price', 'net_income', 'coupon') else \
                     ('center' if cid in ('status', 'refund') else 'w')
            ot.column(cid, width=w, minwidth=40, anchor=anchor)

        ot.tag_configure('refunded',  foreground=RED)
        ot.tag_configure('partial',   foreground='#B05A00')
        ot.tag_configure('cancelled', foreground=GRAY)

        hsb = ttk.Scrollbar(oc, orient='horizontal', command=ot.xview)
        ot.configure(xscrollcommand=hsb.set)
        ot.pack(fill='x')
        hsb.pack(fill='x')

        order_by_iid  = {}
        product_tips  = {}  # iid → tooltip text for product column

        for o in os_:
            has_refund   = bool((o['refund_status'] or '').strip())
            is_cancelled = (o['order_status'] or '') == '不成立'
            note_text    = o['buyer_note'] or o['seller_note'] or ''
            orig         = o['original_price'] or 0
            refunded     = _get_refunded(o)

            coupon_val = o['seller_coupon'] or 0.0
            coupon_str = f'NT$ {coupon_val:,.0f}' if coupon_val else '－'

            # orig_str: show effective price; annotate partial refund
            if refunded > 0 and refunded < orig:
                orig_str = f"NT$ {orig - refunded:,.0f}（退{refunded:,.0f}）"
            else:
                orig_str = f"NT$ {orig:,.0f}"

            if is_cancelled:
                net_str = '－'
            else:
                net = _calc_net(o)
                net_str = f'-NT$ {abs(net):,.0f}' if net < 0 else f'NT$ {net:,.0f}'

            # Multi-item display
            prod_cell, prod_tip = App._product_display(o['product_name'])

            if has_refund and refunded > 0 and refunded < orig:
                tags = ('partial',)   # partial refund: orange
            elif has_refund:
                tags = ('refunded',)  # full refund: red
            elif is_cancelled:
                tags = ('cancelled',)
            else:
                tags = ()
            iid  = ot.insert('', 'end', tags=tags, values=(
                o['order_id'],
                App._short_status(o['order_status']),
                o['refund_status'] or '－',
                (o['order_date'] or '')[:16],
                (o['order_complete_time'] or '')[:16],
                prod_cell,
                orig_str,
                net_str,
                coupon_str,
                o['payment_method'],
                o['shipping_method'],
                o['tracking_number'],
                o['city'],
                o['district'],
                note_text,
            ))
            order_by_iid[iid] = o
            if prod_tip:
                product_tips[iid] = prod_tip

        ot.bind('<Double-1>',
                lambda e, t=ot, d=order_by_iid: self._order_detail_popup(t, d))

        # Custom tooltip: shows full product list for multi-item orders
        def _order_tip(e, t=ot, tips=product_tips, cols=ocols):
            item = t.identify_row(e.y)
            col  = t.identify_column(e.x)
            if not item or not col:
                return ''
            col_idx = int(col[1:]) - 1
            if col_idx < len(cols) and cols[col_idx] == 'product':
                if item in tips:
                    return tips[item]
            # Default: show full text if truncated
            vals = t.item(item, 'values')
            if col_idx < len(vals):
                text = str(vals[col_idx])
                return text if len(text) > 20 else ''
            return ''

        _Tooltip(ot, _order_tip)
        self._cv.yview_moveto(0)

    def _order_detail_popup(self, tree, order_by_iid):
        sel = tree.selection()
        if not sel:
            return
        o = order_by_iid.get(sel[0])
        if not o:
            return

        ni = _calc_net(o)

        # Parse product list for multi-item display
        raw_product = o['product_name'] or ''
        product_parts = [s.strip() for s in raw_product.split('|') if s.strip()]
        multi_item    = len(product_parts) > 1

        dlg = tk.Toplevel(self)
        dlg.title('訂單詳情')
        win_h = 680 if multi_item and len(product_parts) > 2 else 640
        dlg.geometry(f'580x{win_h}')
        dlg.configure(bg=WHITE)
        dlg.transient(self)
        dlg.grab_set()

        tk.Label(dlg, text=f'訂單 {o["order_id"]}', bg=WHITE, fg=JADE,
                 font=(FONT, 12, 'bold')).pack(anchor='w', padx=20, pady=(14, 6))
        ttk.Separator(dlg, orient='horizontal').pack(fill='x', padx=20, pady=(0, 8))

        # ── Scrollable content ─────────────────────────────────────────────
        wrapper  = tk.Frame(dlg, bg=WHITE)
        wrapper.pack(fill='both', expand=True, padx=20)
        cv   = tk.Canvas(wrapper, bg=WHITE, highlightthickness=0)
        vsb  = ttk.Scrollbar(wrapper, orient='vertical', command=cv.yview)
        cv.configure(yscrollcommand=vsb.set)
        vsb.pack(side='right', fill='y')
        cv.pack(side='left', fill='both', expand=True)
        fr = tk.Frame(cv, bg=WHITE)
        fr_id = cv.create_window((0, 0), window=fr, anchor='nw')
        fr.bind('<Configure>', lambda _: cv.configure(scrollregion=cv.bbox('all')))
        cv.bind('<Configure>', lambda e: cv.itemconfigure(fr_id, width=e.width))
        def _mwheel(e): cv.yview_scroll(-1*(e.delta//120), 'units')
        cv.bind('<Enter>', lambda _: cv.bind_all('<MouseWheel>', _mwheel))
        cv.bind('<Leave>', lambda _: cv.unbind_all('<MouseWheel>'))

        fr.columnconfigure(1, weight=1)
        row_i = 0

        def add_field(label, value, height=1):
            nonlocal row_i
            tk.Label(fr, text=label, bg=WHITE, fg=GRAY,
                     font=(FONT, 9), anchor='ne' if height > 1 else 'e').grid(
                row=row_i, column=0, sticky='ne' if height > 1 else 'e',
                padx=(8, 8), pady=2)
            if height == 1:
                e = tk.Entry(fr, font=(FONT, 10), relief='solid', bd=1,
                             readonlybackground='#F9F9F9', fg=TEXT)
                e.insert(0, value or '')
                e.config(state='readonly')
                e.grid(row=row_i, column=1, sticky='ew', pady=2)
            else:
                t = tk.Text(fr, font=(FONT, 10), relief='solid', bd=1,
                            bg='#F9F9F9', fg=TEXT, height=height, wrap='word')
                t.insert('1.0', value or '')
                t.config(state='disabled')
                t.grid(row=row_i, column=1, sticky='ew', pady=2)
            row_i += 1

        add_field('訂單編號',     o['order_id'])
        add_field('訂單狀態',     o['order_status'])
        add_field('退貨退款狀態', o['refund_status'])
        add_field('訂單成立日期', o['order_date'])
        add_field('訂單完成時間', o['order_complete_time'])

        # Product area
        if multi_item:
            product_text = '\n'.join(f'{i+1}. {p}' for i, p in enumerate(product_parts))
            add_field('商品明細', product_text, height=min(len(product_parts) + 1, 5))
        else:
            add_field('商品名稱', raw_product)
            if (o['product_option'] or '').strip():
                add_field('商品選項', o['product_option'])

        popup_orig     = o['original_price'] or 0
        popup_refunded = _get_refunded(o)
        add_field('數量',         str(o['quantity']))
        add_field('商品原價',     f"NT$ {popup_orig:,.0f}")
        if popup_refunded > 0:
            label_r = '退款（部分）' if popup_refunded < popup_orig else '退款（全退）'
            add_field(label_r,    f"NT$ {popup_refunded:,.0f}")
        add_field('賣家優惠券',   f"NT$ {(o['seller_coupon'] or 0):,.0f}")
        add_field('成交手續費',   f"NT$ {(o['transaction_fee'] or 0):,.0f}")
        add_field('其他服務費',   f"NT$ {(o['service_fee'] or 0):,.0f}")
        add_field('金流處理費',   f"NT$ {(o['payment_fee'] or 0):,.0f}")
        add_field('退貨運費',     f"NT$ {(o['return_shipping'] or 0):,.0f}")
        add_field('進帳金額',     (f'-NT$ {abs(ni):,.0f}' if ni < 0 else f'NT$ {ni:,.0f}'))
        add_field('付款方式',     o['payment_method'])
        add_field('寄送方式',     o['shipping_method'])
        add_field('包裹查詢號碼', o['tracking_number'])
        add_field('城市',         o['city'])
        add_field('行政區',       o['district'])
        add_field('買家備註',     o['buyer_note'])
        add_field('賣家備註',     o['seller_note'])

        tk.Label(dlg, text='（欄位可點選後 Ctrl+A → Ctrl+C 複製）',
                 bg=WHITE, fg='#BBB', font=(FONT, 8)).pack(pady=(4, 2))
        self._btn(dlg, '關閉', GRAY, dlg.destroy).pack(pady=(2, 12))

    # ── Delete / Undo ────────────────────────────────────────────────────────

    def _delete_customer(self):
        if not self._current:
            return
        acct = self._current
        conn = get_db()
        c   = conn.execute('SELECT * FROM customers WHERE shopee_account=?', (acct,)).fetchone()
        os_ = conn.execute('SELECT * FROM orders   WHERE shopee_account=?', (acct,)).fetchall()
        conn.close()
        if not c:
            return
        name_display = c['real_name'] or acct
        order_count  = len(os_)
        confirmed = messagebox.askyesno(
            '確認刪除客戶',
            f'確定要刪除客戶「{name_display}」嗎？\n\n'
            f'蝦皮帳號：{acct}\n'
            f'相關訂單：{order_count} 筆\n\n'
            f'⚠ 刪除後 10 秒內可在畫面下方點「復原」取消。\n'
            f'（超過 10 秒後將無法復原）',
            parent=self)
        if not confirmed:
            return
        backup = {'customer': dict(c), 'orders': [dict(o) for o in os_]}
        conn = get_db()
        conn.execute('DELETE FROM orders   WHERE shopee_account=?', (acct,))
        conn.execute('DELETE FROM customers WHERE shopee_account=?', (acct,))
        conn.commit()
        conn.close()
        self._current = None
        self._show_placeholder()
        self._refresh_list()
        self.status_var.set(f'已刪除：{name_display}（{order_count} 筆訂單）')
        self._start_undo_countdown(backup)

    def _start_undo_countdown(self, backup, seconds=10):
        if self._undo_job:
            self.after_cancel(self._undo_job)
        self._undo_backup  = backup
        self._undo_seconds = seconds
        self._undo_btn.pack(side='right', padx=(0, 8))
        self._tick_undo()

    def _tick_undo(self):
        if self._undo_seconds <= 0:
            self._undo_btn.pack_forget()
            self._undo_backup = None
            self._undo_job    = None
            return
        self._undo_btn.config(text=f'↩ 復原刪除（{self._undo_seconds} 秒）')
        self._undo_seconds -= 1
        self._undo_job = self.after(1000, self._tick_undo)

    def _undo_delete(self):
        if not self._undo_backup:
            return
        backup = self._undo_backup
        if self._undo_job:
            self.after_cancel(self._undo_job)
            self._undo_job = None
        self._undo_btn.pack_forget()
        self._undo_backup = None
        c   = backup['customer']
        os_ = backup['orders']
        conn = get_db()
        try:
            conn.execute(
                'INSERT INTO customers (shopee_account,real_name,line_account,birthday,'
                'preferences,notes,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?)',
                (c['shopee_account'], c['real_name'], c['line_account'], c['birthday'],
                 c['preferences'], c['notes'], c['created_at'], c['updated_at']))
        except sqlite3.IntegrityError:
            pass
        for o in os_:
            try:
                conn.execute(
                    'INSERT INTO orders (order_id,shopee_account,order_status,refund_status,'
                    'order_date,product_name,product_option,quantity,product_price,paid_amount,'
                    'original_price,seller_coupon,transaction_fee,service_fee,payment_fee,'
                    'return_shipping,city,district,payment_method,shipping_method,'
                    'tracking_number,order_complete_time,buyer_note,seller_note) VALUES '
                    '(:order_id,:shopee_account,:order_status,:refund_status,'
                    ':order_date,:product_name,:product_option,:quantity,:product_price,:paid_amount,'
                    ':original_price,:seller_coupon,:transaction_fee,:service_fee,:payment_fee,'
                    ':return_shipping,:city,:district,:payment_method,:shipping_method,'
                    ':tracking_number,:order_complete_time,:buyer_note,:seller_note)', o)
            except sqlite3.IntegrityError:
                pass
        conn.commit()
        conn.close()
        acct = c['shopee_account']
        self._refresh_list()
        if acct in self.tree.get_children():
            self.tree.selection_set(acct)
            self.tree.see(acct)
        self._current = acct
        self._show_detail(acct)
        self.status_var.set(f'已復原：{c["real_name"] or acct}')

    # ── Data operations ─────────────────────────────────────────────────────

    def _refresh_list(self):
        start_d = self.start_date_var.get().strip()
        end_d   = self.end_date_var.get().strip()

        # Build date clause pieces for order-level queries
        d_cond   = ''
        d_params = []
        if start_d:
            d_cond += ' AND order_date >= ?'; d_params.append(start_d)
        if end_d:
            d_cond += ' AND order_date <= ?'; d_params.append(end_d + ' 23:59:59')

        # ── Sort ──────────────────────────────────────────────────────────
        if self._sort_col:
            _col_sql = {
                'acct': 'c.shopee_account', 'name': 'c.real_name',
                'line': 'c.line_account',   'cnt':  'cnt',
                'spent': 'rev',             'last': 'last_date',
            }
            d = 'DESC' if self._sort_dir == 'desc' else 'ASC'
            if self._sort_col == 'name':
                sort = (f'ORDER BY CASE WHEN c.real_name="" OR c.real_name IS NULL '
                        f'THEN 1 ELSE 0 END, c.real_name {d}')
            else:
                sort = f'ORDER BY {_col_sql[self._sort_col]} {d}'
        else:
            sort = {
                '最近訂購': 'ORDER BY last_date DESC',
                '消費金額': 'ORDER BY rev DESC',
                '訂單數量': 'ORDER BY cnt DESC',
                '姓名':    ('ORDER BY CASE WHEN c.real_name="" OR c.real_name IS NULL '
                            'THEN 1 ELSE 0 END, c.real_name'),
            }.get(self.sort_var.get(), 'ORDER BY last_date DESC')

        # ── Main list query ───────────────────────────────────────────────
        q = self.search_var.get().strip()
        where_parts = []
        params      = []
        if q:
            where_parts.append(
                '(c.shopee_account LIKE ? OR c.real_name LIKE ? OR c.line_account LIKE ?)')
            params += [f'%{q}%'] * 3
        if start_d:
            where_parts.append('o.order_date >= ?'); params.append(start_d)
        if end_d:
            where_parts.append('o.order_date <= ?'); params.append(end_d + ' 23:59:59')

        where_clause = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''

        sql = (f'SELECT c.shopee_account, c.real_name, c.line_account,'
               f'COUNT(CASE WHEN o.order_status!="不成立" AND '
               f'     (o.refund_status="" OR o.refund_status IS NULL '
               f'      OR (COALESCE(o.refunded_price,0)>0 AND o.refunded_price<o.original_price)) THEN 1 END) cnt,'
               f'COALESCE(SUM(CASE WHEN o.order_status!="不成立" '
               f'    AND (COALESCE(o.refunded_price,0)>0 OR o.refund_status="" OR o.refund_status IS NULL) '
               f'    THEN (o.original_price - COALESCE(o.refunded_price,0)) ELSE 0 END),0) sales_rev,'
               f'COALESCE(SUM(CASE WHEN o.order_status!="不成立" THEN {_NET_SQL} ELSE 0 END),0) rev,'
               f'MAX(o.order_date) last_date '
               f'FROM customers c LEFT JOIN orders o ON c.shopee_account=o.shopee_account '
               f'{where_clause} GROUP BY c.id {sort}')

        conn = get_db()
        rows = conn.execute(sql, params).fetchall()

        # ── Header stats (respect date filter) ───────────────────────────
        if start_d or end_d:
            tc = conn.execute(
                f'SELECT COUNT(DISTINCT shopee_account) FROM orders WHERE 1=1{d_cond}',
                d_params).fetchone()[0]
        else:
            tc = conn.execute('SELECT COUNT(*) FROM customers').fetchone()[0]

        ta = conn.execute(
            f"SELECT COUNT(*) FROM orders "
            f"WHERE order_status!='不成立' "
            f"AND (refund_status='' OR refund_status IS NULL "
            f"     OR (COALESCE(refunded_price,0)>0 AND refunded_price<original_price)){d_cond}",
            d_params).fetchone()[0]
        refund_cnt = conn.execute(
            f"SELECT COUNT(*) FROM orders "
            f"WHERE refund_status!='' AND refund_status IS NOT NULL{d_cond}",
            d_params).fetchone()[0]
        ts = conn.execute(
            f"SELECT COALESCE(SUM(CASE "
            f"  WHEN COALESCE(refunded_price,0)>0 OR refund_status='' OR refund_status IS NULL "
            f"  THEN (original_price - COALESCE(refunded_price,0)) ELSE 0 END),0) "
            f"FROM orders WHERE order_status!='不成立'{d_cond}",
            d_params).fetchone()[0]
        tr = conn.execute(
            f"SELECT COALESCE(SUM({_NET_SQL}),0) FROM orders "
            f"WHERE order_status!='不成立'{d_cond}",
            d_params).fetchone()[0]
        conn.close()

        self.lbl_cust.config(text=f'客戶：{tc}')
        self.lbl_all.config(text=f'訂單：{ta}')
        self.lbl_refund.config(text=f'已退貨：{refund_cnt}')
        self.lbl_sales.config(text=f'銷售總額：NT$ {ts:,.0f}')
        self.lbl_rev.config(text=f'進帳總額：NT$ {tr:,.0f}')

        self.tree.delete(*self.tree.get_children())
        for r in rows:
            acct  = r['shopee_account']
            name  = ('★ ' if (r['cnt'] >= 5 and r['sales_rev'] >= 10000) else '') + (r['real_name'] or '')
            spent = f"NT$ {r['rev']:,.0f}" if r['rev'] else '-'
            last  = (r['last_date'] or '')[:10]
            self.tree.insert('', 'end', iid=acct,
                             values=(acct, name, r['line_account'] or '', r['cnt'], spent, last))

        # Date filter indicator
        if start_d or end_d:
            self._filter_lbl.config(text=f'● {start_d or "─"} ～ {end_d or "─"}')
            if not self._clear_date_btn.winfo_ismapped():
                self._clear_date_btn.pack(side='left', padx=(4, 0))
            status_extra = f' ／ 期間：{start_d or "─"}～{end_d or "─"}'
        else:
            self._filter_lbl.config(text='')
            self._clear_date_btn.pack_forget()
            status_extra = ''
        self.status_var.set(f'共 {len(rows)} 位客戶{status_extra}')
        self._update_heading_arrows()

    def _on_select(self, _event):
        sel = self.tree.selection()
        if not sel:
            return
        self._current = sel[0]
        self._show_detail(self._current)

    def _copy_to_clipboard(self, text):
        self.clipboard_clear()
        self.clipboard_append(text)
        self.status_var.set(f'已複製：{text}')

    def _on_list_rightclick(self, e):
        item = self.tree.identify_row(e.y)
        if not item:
            return
        self.tree.selection_set(item)
        vals = self.tree.item(item, 'values')
        acct = vals[0] if vals else ''
        name = vals[1].lstrip('★').strip() if vals else ''
        line = vals[2] if vals else ''
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label=f'複製蝦皮帳號　{acct}',
                         command=lambda: self._copy_to_clipboard(acct))
        if name:
            menu.add_command(label=f'複製姓名　{name}',
                             command=lambda n=name: self._copy_to_clipboard(n))
        if line:
            menu.add_command(label=f'複製 LINE　{line}',
                             command=lambda l=line: self._copy_to_clipboard(l))
        menu.tk_popup(e.x_root, e.y_root)

    def _save(self):
        if not self._current:
            return
        conn = get_db()
        conn.execute(
            'UPDATE customers SET real_name=?,line_account=?,birthday=?,'
            'preferences=?,notes=?,updated_at=datetime("now","localtime") '
            'WHERE shopee_account=?',
            (self._field_vars['real_name'].get().strip(),
             self._field_vars['line_account'].get().strip(),
             self._field_vars['birthday'].get().strip(),
             self._pref_text.get('1.0', 'end-1c').strip(),
             self._notes_text.get('1.0', 'end-1c').strip(),
             self._current))
        conn.commit()
        conn.close()
        self.status_var.set('客戶資料已更新')
        self._refresh_list()
        if self._current in self.tree.get_children():
            self.tree.selection_set(self._current)
        self._show_detail(self._current)

    def _clear_date_filter(self):
        self.start_date_var.set('')
        self.end_date_var.set('')

    def _open_date_picker(self):
        _DateRangePicker(self, self.start_date_var, self.end_date_var)

    def _open_chart(self):
        _ChartWindow(self,
                     start_date=self.start_date_var.get().strip() or None,
                     end_date=self.end_date_var.get().strip() or None)

    def _import_excel(self):
        paths = filedialog.askopenfilenames(
            title='選擇蝦皮訂單報表（可多選）',
            filetypes=[('Excel 檔案', '*.xlsx *.xls'), ('所有檔案', '*.*')],
            parent=self)
        if not paths:
            return

        total = len(paths)
        self.status_var.set(f'讀取 {total} 個檔案中...')
        self.update()

        # Phase 1: parse all files
        parsed = []
        for path in paths:
            try:
                order_map, price_acc, filename = _parse_excel(path)
                dups = _find_dup_orders(order_map)
                parsed.append((order_map, price_acc, filename, dups))
            except Exception as e:
                messagebox.showerror(
                    '讀取失敗',
                    f'無法讀取檔案：\n{os.path.basename(path)}\n\n{e}\n\n'
                    '請確認檔案已在 Excel 中關閉，再重試。',
                    parent=self)
                self.status_var.set('匯入失敗')
                return

        # Phase 2: check duplicates globally, ask once
        all_dups = [oid for _, _, _, dups in parsed for oid in dups]
        skip_existing = False
        if all_dups:
            n      = len(all_dups)
            sample = '\n'.join(f'  • {oid}' for oid in all_dups[:15])
            if n > 15:
                sample += f'\n  （還有 {n - 15} 筆…）'
            answer = messagebox.askyesno(
                '發現重複訂單',
                f'在 {total} 個檔案中共發現 {n} 筆重複訂單：\n\n{sample}\n\n'
                f'是否覆蓋更新這些訂單的資料？\n\n'
                f'【注意】系統會自動比對訂單完成時間，\n'
                f'若系統內已有較新的資料，該筆訂單會自動跳過，\n'
                f'不會被舊資料覆蓋。\n\n'
                f'（選「否」則略過所有重複訂單，只匯入全新資料）',
                parent=self)
            skip_existing = not answer

        # Phase 3: show progress dialog and import
        prog = self._make_excel_progress(total)
        total_nc = total_no = total_uo = total_sk = 0
        errors = []

        for i, (order_map, price_acc, filename, dups) in enumerate(parsed):
            prog.advance(i, filename)
            try:
                nc, no, uo = _write_to_db(order_map, price_acc, filename, skip_existing)
                sk = len(dups) if skip_existing else 0
                total_nc += nc; total_no += no; total_uo += uo; total_sk += sk
                prog.mark_done(filename, True, nc, no, uo, sk)
            except Exception as e:
                errors.append((filename, str(e)))
                prog.mark_done(filename, False, 0, 0, 0, 0)

        prog.close()
        _scan_line_from_notes()
        self._refresh_list()

        if errors:
            err_text = '\n'.join(f'  • {fn}：{msg}' for fn, msg in errors)
            messagebox.showwarning(
                '部分失敗',
                f'以下 {len(errors)} 個檔案寫入失敗：\n\n{err_text}',
                parent=self)

        msg = (f'匯入完成！共處理 {total} 個檔案。\n\n'
               f'新增客戶：{total_nc} 位\n新增訂單：{total_no} 筆\n更新訂單：{total_uo} 筆')
        if total_sk:
            msg += f'\n略過重複：{total_sk} 筆'
        if errors:
            msg += f'\n失敗檔案：{len(errors)} 個'
        messagebox.showinfo('匯入完成', msg, parent=self)
        self.status_var.set(
            f'匯入完成：+{total_nc} 位客戶，+{total_no} 筆新訂單，更新 {total_uo} 筆')

    def _make_excel_progress(self, total: int):
        """Progress dialog for Excel batch import."""
        dlg = tk.Toplevel(self)
        dlg.title('匯入報表中...')
        dlg.geometry('480x230')
        dlg.resizable(False, False)
        dlg.configure(bg=WHITE)
        dlg.transient(self)
        dlg.protocol('WM_DELETE_WINDOW', lambda: None)

        tk.Label(dlg, text='匯入報表中，請稍候…', bg=WHITE, fg=JADE,
                 font=(FONT, 12, 'bold')).pack(pady=(18, 6), padx=20, anchor='w')

        file_lbl = tk.Label(dlg, text='準備中…', bg=WHITE, fg=TEXT,
                            font=(FONT, 9), anchor='w', wraplength=440)
        file_lbl.pack(padx=20, anchor='w')

        bar_frame = tk.Frame(dlg, bg=WHITE)
        bar_frame.pack(fill='x', padx=20, pady=(10, 4))
        bar = ttk.Progressbar(bar_frame, length=440, mode='determinate', maximum=total)
        bar.pack(fill='x')

        bottom = tk.Frame(dlg, bg=WHITE)
        bottom.pack(fill='x', padx=20)
        count_lbl = tk.Label(bottom, text=f'0 / {total} 個', bg=WHITE, fg=GRAY, font=(FONT, 9))
        count_lbl.pack(side='left')
        status_lbl = tk.Label(bottom, text='', bg=WHITE, font=(FONT, 9))
        status_lbl.pack(side='right')

        log_frame = tk.Frame(dlg, bg=WHITE)
        log_frame.pack(fill='x', padx=20, pady=(6, 0))
        log_lbl = tk.Label(log_frame, text='', bg=WHITE, fg=GRAY,
                           font=(FONT, 8), anchor='w', wraplength=440, justify='left')
        log_lbl.pack(anchor='w')

        dlg.update()
        log_lines = []

        class _Handle:
            def advance(self_, idx, fname):
                file_lbl.config(text=f'正在處理：{fname}')
                bar['value'] = idx
                count_lbl.config(text=f'{idx} / {total} 個')
                status_lbl.config(text='處理中…', fg='#B07000')
                dlg.update()

            def mark_done(self_, fname, ok: bool, nc, no, uo, sk):
                bar['value'] = bar['value'] + 1
                count_lbl.config(text=f'{int(bar["value"])} / {total} 個')
                if ok:
                    detail = f'+{nc}客戶 +{no}訂單 更新{uo}'
                    if sk:
                        detail += f' 略過{sk}'
                    log_lines.append(('✓', fname, detail, JADE))
                    status_lbl.config(text='完成', fg=JADE)
                else:
                    log_lines.append(('✕', fname, '寫入失敗', RED))
                    status_lbl.config(text='失敗', fg=RED)
                shown = log_lines[-3:]
                log_lbl.config(
                    text='\n'.join(f'{ic} {fn}  {det}' for ic, fn, det, _ in shown),
                    fg=shown[-1][3])
                dlg.update()

            def close(self_):
                if dlg.winfo_exists():
                    dlg.destroy()

        return _Handle()

    def _make_ocr_progress(self, total: int):
        """Create a non-blocking progress dialog for OCR. Returns a handle with
        .advance(i, fname), .mark_done(fname, ok), and .close() methods."""
        dlg = tk.Toplevel(self)
        dlg.title('辨識圖片中...')
        dlg.geometry('480x210')
        dlg.resizable(False, False)
        dlg.configure(bg=WHITE)
        dlg.transient(self)
        dlg.protocol('WM_DELETE_WINDOW', lambda: None)  # block close button

        tk.Label(dlg, text='AI 圖片辨識中，請稍候…', bg=WHITE, fg=JADE,
                 font=(FONT, 12, 'bold')).pack(pady=(18, 6), padx=20, anchor='w')

        file_lbl = tk.Label(dlg, text='準備中…', bg=WHITE, fg=TEXT,
                            font=(FONT, 9), anchor='w', wraplength=440)
        file_lbl.pack(padx=20, anchor='w')

        bar_frame = tk.Frame(dlg, bg=WHITE)
        bar_frame.pack(fill='x', padx=20, pady=(10, 4))
        bar = ttk.Progressbar(bar_frame, length=440, mode='determinate', maximum=total)
        bar.pack(fill='x')

        bottom = tk.Frame(dlg, bg=WHITE)
        bottom.pack(fill='x', padx=20)
        count_lbl = tk.Label(bottom, text=f'0 / {total} 張', bg=WHITE, fg=GRAY,
                             font=(FONT, 9))
        count_lbl.pack(side='left')
        status_lbl = tk.Label(bottom, text='', bg=WHITE, font=(FONT, 9))
        status_lbl.pack(side='right')

        log_frame = tk.Frame(dlg, bg=WHITE)
        log_frame.pack(fill='x', padx=20, pady=(6, 0))
        log_lbl = tk.Label(log_frame, text='', bg=WHITE, fg=GRAY,
                           font=(FONT, 8), anchor='w', wraplength=440, justify='left')
        log_lbl.pack(anchor='w')

        dlg.update()
        log_lines = []

        class _Handle:
            def advance(self_, idx, fname):
                file_lbl.config(text=f'正在辨識：{fname}')
                bar['value'] = idx
                count_lbl.config(text=f'{idx} / {total} 張')
                status_lbl.config(text='辨識中…', fg='#B07000')
                dlg.update()

            def mark_done(self_, fname, ok: bool):
                icon = '✓' if ok else '✕'
                fg   = JADE if ok else RED
                log_lines.append((icon, fname, fg))
                bar['value'] = bar['value'] + 1
                count_lbl.config(text=f'{int(bar["value"])} / {total} 張')
                status_lbl.config(text='完成' if ok else '失敗', fg=fg)
                shown = log_lines[-3:]
                log_lbl.config(
                    text='\n'.join(f'{ic} {fn}' for ic, fn, _ in shown),
                    fg=shown[-1][2])
                dlg.update()

            def close(self_):
                if dlg.winfo_exists():
                    dlg.destroy()

        return _Handle()

    def _show_api_key_dialog(self) -> bool:
        """Show API key setup dialog. Returns True if a key was saved."""
        dlg = tk.Toplevel(self)
        dlg.title('設定 API 金鑰')
        dlg.geometry('520x430')
        dlg.resizable(False, False)
        dlg.configure(bg=WHITE)
        dlg.grab_set()
        dlg.transient(self)

        settings = _load_settings()

        tk.Label(dlg, text='設定圖片辨識 API 金鑰', bg=WHITE, fg=JADE,
                 font=(FONT, 13, 'bold')).pack(pady=(18, 2), padx=24, anchor='w')

        # ── Gemini（免費）區塊 ──
        gem_frame = tk.LabelFrame(dlg, text='★ Google Gemini（免費推薦）',
                                  bg=WHITE, fg='#2A6C28', font=(FONT, 10, 'bold'),
                                  relief='groove', bd=2)
        gem_frame.pack(fill='x', padx=20, pady=(8, 4))

        tk.Label(gem_frame, text=(
            '申請：aistudio.google.com → 左側「Get API key」→「建立 API 金鑰」\n'
            '金鑰格式以 AIza 開頭。每天免費 1500 次，無需信用卡。'
        ), bg=WHITE, fg=TEXT, font=(FONT, 9), justify='left').pack(padx=10, pady=(4, 2), anchor='w')

        gef = tk.Frame(gem_frame, bg=WHITE)
        gef.pack(fill='x', padx=10, pady=(0, 8))
        tk.Label(gef, text='Gemini Key：', bg=WHITE, font=(FONT, 10), width=12, anchor='e').pack(side='left')
        gem_var = tk.StringVar(value=settings.get('gemini_api_key', ''))
        gem_entry = tk.Entry(gef, textvariable=gem_var, font=(FONT, 10),
                             width=32, show='*', relief='solid', bd=1)
        gem_entry.pack(side='left', padx=(4, 4))
        def do_clear_gem():
            if messagebox.askyesno('確認清除', '確定要清除 Gemini 金鑰嗎？', parent=dlg):
                s = _load_settings(); s.pop('gemini_api_key', None); _save_settings(s)
                gem_var.set('')
        tk.Button(gef, text='清除', font=(FONT, 9), bg='#EEEEEE', relief='flat',
                  command=do_clear_gem).pack(side='left')

        # ── Anthropic（付費）區塊 ──
        ant_frame = tk.LabelFrame(dlg, text='Anthropic（付費備選）',
                                  bg=WHITE, fg=GRAY, font=(FONT, 10),
                                  relief='groove', bd=2)
        ant_frame.pack(fill='x', padx=20, pady=(4, 4))

        tk.Label(ant_frame, text=(
            '申請：console.anthropic.com → 左側「API Keys」→「Create Key」\n'
            '金鑰格式以 sk-ant- 開頭。需付費，若已設定 Gemini 則不需要。'
        ), bg=WHITE, fg=GRAY, font=(FONT, 9), justify='left').pack(padx=10, pady=(4, 2), anchor='w')

        anf = tk.Frame(ant_frame, bg=WHITE)
        anf.pack(fill='x', padx=10, pady=(0, 8))
        tk.Label(anf, text='Anthropic Key：', bg=WHITE, font=(FONT, 10), width=12, anchor='e').pack(side='left')
        ant_var = tk.StringVar(value=settings.get('anthropic_api_key', ''))
        ant_entry = tk.Entry(anf, textvariable=ant_var, font=(FONT, 10),
                             width=32, show='*', relief='solid', bd=1)
        ant_entry.pack(side='left', padx=(4, 4))
        def do_clear_ant():
            if messagebox.askyesno('確認清除', '確定要清除 Anthropic 金鑰嗎？', parent=dlg):
                s = _load_settings(); s.pop('anthropic_api_key', None); _save_settings(s)
                ant_var.set('')
        tk.Button(anf, text='清除', font=(FONT, 9), bg='#EEEEEE', relief='flat',
                  command=do_clear_ant).pack(side='left')

        tk.Label(dlg, text='若同時填入兩組金鑰，優先使用 Gemini（免費）。',
                 bg=WHITE, fg=GRAY, font=(FONT, 9)).pack(pady=(2, 4))

        saved = [False]

        def do_save():
            gk = gem_var.get().strip()
            ak = ant_var.get().strip()
            if not gk and not ak:
                messagebox.showwarning('請輸入金鑰', '至少需要填入一組 API 金鑰。', parent=dlg)
                return
            s = _load_settings()
            if gk:
                s['gemini_api_key'] = gk
            else:
                s.pop('gemini_api_key', None)
            if ak:
                s['anthropic_api_key'] = ak
            else:
                s.pop('anthropic_api_key', None)
            _save_settings(s)
            saved[0] = True
            dlg.destroy()

        bf = tk.Frame(dlg, bg=WHITE)
        bf.pack(pady=(4, 0))
        self._btn(bf, '儲存', JADE, do_save).pack(side='left', padx=8)
        self._btn(bf, '取消', '#888888', dlg.destroy).pack(side='left', padx=8)

        gem_entry.focus()
        dlg.wait_window()
        return saved[0]

    def _ensure_api_key(self) -> bool:
        """Check auth availability; show setup dialog if missing. Returns True if ready."""
        if _get_api_auth() is not None:
            return True
        return self._show_api_key_dialog()

    def _import_images(self):
        paths = filedialog.askopenfilenames(
            title='選擇出貨標籤圖片（可多選）',
            filetypes=[('圖片檔案', '*.jpg *.jpeg *.png *.webp'), ('所有檔案', '*.*')],
            parent=self)
        if not paths:
            return

        if not self._ensure_api_key():
            return

        total = len(paths)
        # pending: rows that OCR'd successfully and have a matched account
        # Each item: {'fname', 'order_id', 'acct', 'ocr_name', 'existing_name'}
        pending = []
        errors  = []   # (fname, reason)

        # ── Phase 1: OCR all images (with progress dialog) ───────────────
        prog = self._make_ocr_progress(total)

        for i, path in enumerate(paths):
            fname = os.path.basename(path)
            prog.advance(i, fname)

            try:
                ocr = _ocr_shipping_label(path)
            except RuntimeError as e:
                msg = str(e)
                if msg in ('AUTH_MISSING', 'AUTH_INVALID'):
                    prog.close()
                    label = '金鑰無效或已過期' if msg == 'AUTH_INVALID' else '未設定金鑰'
                    retry = messagebox.askyesno(
                        f'API 金鑰問題（{label}）',
                        f'圖片「{fname}」辨識失敗：{label}。\n\n是否立即設定 API 金鑰並重試？',
                        parent=self)
                    if retry and self._show_api_key_dialog():
                        prog = self._make_ocr_progress(total)
                        prog.advance(i, fname)
                        try:
                            ocr = _ocr_shipping_label(path)
                        except Exception as e2:
                            errors.append((fname, str(e2)))
                            prog.mark_done(fname, ok=False)
                            continue
                    else:
                        prog.close()
                        self.status_var.set('已取消')
                        return
                else:
                    errors.append((fname, msg))
                    prog.mark_done(fname, ok=False)
                    continue
            except Exception as e:
                errors.append((fname, str(e)))
                prog.mark_done(fname, ok=False)
                continue

            order_id  = (ocr.get('order_id')  or '').strip()
            real_name = (ocr.get('real_name') or '').strip()

            if not order_id and not real_name:
                errors.append((fname, '圖片中未找到訂單編號或收件人姓名'))
                prog.mark_done(fname, ok=False)
                continue
            if not order_id:
                errors.append((fname, f'找到姓名「{real_name}」但無訂單編號，無法對應帳號'))
                prog.mark_done(fname, ok=False)
                continue
            if not real_name:
                errors.append((fname, f'找到訂單編號「{order_id}」但未找到收件人姓名'))
                prog.mark_done(fname, ok=False)
                continue

            conn = get_db()
            order_row = conn.execute(
                'SELECT shopee_account FROM orders WHERE order_id=?', (order_id,)
            ).fetchone()
            if not order_row:
                conn.close()
                errors.append((fname, f'訂單 {order_id} 不在系統中，請先匯入訂單報表'))
                prog.mark_done(fname, ok=False)
                continue
            acct = order_row['shopee_account']
            cust = conn.execute(
                'SELECT real_name FROM customers WHERE shopee_account=?', (acct,)
            ).fetchone()
            existing = (cust['real_name'] if cust else '') or ''
            conn.close()

            pending.append({
                'fname': fname, 'order_id': order_id,
                'acct': acct, 'ocr_name': real_name, 'existing_name': existing,
            })
            prog.mark_done(fname, ok=True)

        prog.close()

        if not pending and not errors:
            self.status_var.set('無可辨識的圖片')
            return

        # ── Phase 2: Review dialog (edit before write) ────────────────────
        confirmed = [] if not pending else self._show_ocr_review(pending)

        # ── Phase 3: Write confirmed results ──────────────────────────────
        written = skipped = 0
        for item in (confirmed or []):
            conn = get_db()
            conn.execute('UPDATE customers SET real_name=? WHERE shopee_account=?',
                         (item['final_name'], item['acct']))
            conn.commit()
            conn.close()
            written += 1

        if confirmed is not None:
            skipped = len(pending) - written

        self._refresh_list()

        # ── Summary ───────────────────────────────────────────────────────
        lines = []
        if confirmed is not None:
            for item in (confirmed or []):
                tag = '（已更新）' if item['existing_name'] else '（新寫入）'
                lines.append(f'✓  {item["fname"]}\n   {item["acct"]} → {item["final_name"]} {tag}')
            if skipped:
                lines.append(f'○  {skipped} 張略過（使用者取消）')
        for fname, reason in errors:
            lines.append(f'✕  {fname}\n   {reason}')

        body = f'共 {total} 張圖片　寫入 {written} 筆　失敗 {len(errors)} 張\n\n'
        body += '\n'.join(lines)
        if errors:
            messagebox.showwarning('圖片辨識完成', body, parent=self)
        else:
            messagebox.showinfo('圖片辨識完成', body, parent=self)
        self.status_var.set(f'圖片辨識：寫入 {written} 筆，失敗 {len(errors)} 張')

    def _show_ocr_review(self, pending: list) -> list | None:
        """Show an editable review table of OCR results.
        Returns list of confirmed items (with 'final_name'), or None if cancelled."""
        dlg = tk.Toplevel(self)
        dlg.title('確認辨識結果')
        dlg.geometry('780x440')
        dlg.minsize(620, 300)
        dlg.configure(bg=WHITE)
        dlg.grab_set()
        dlg.transient(self)

        tk.Label(dlg,
                 text='以下是辨識結果，請核對姓名（可直接修改）。取消勾選代表略過該筆。',
                 bg=WHITE, fg=TEXT, font=(FONT, 10)).pack(pady=(14, 6), padx=16, anchor='w')

        # ── Scrollable table ──────────────────────────────────────────────
        outer = tk.Frame(dlg, bg=WHITE)
        outer.pack(fill='both', expand=True, padx=16, pady=0)

        cv = tk.Canvas(outer, bg=WHITE, highlightthickness=0)
        sb = ttk.Scrollbar(outer, orient='vertical', command=cv.yview)
        fr = tk.Frame(cv, bg=WHITE)
        fr_id = cv.create_window((0, 0), window=fr, anchor='nw')
        cv.configure(yscrollcommand=sb.set)
        cv.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')
        fr.bind('<Configure>', lambda _: cv.configure(scrollregion=cv.bbox('all')))
        cv.bind('<Configure>', lambda e: cv.itemconfigure(fr_id, width=e.width))
        cv.bind('<Enter>', lambda _: cv.bind_all('<MouseWheel>',
                lambda e: cv.yview_scroll(-1*(e.delta//120), 'units')))
        cv.bind('<Leave>', lambda _: cv.unbind_all('<MouseWheel>'))

        HDR_BG = '#EAF4EE'
        cols = ['', '圖片檔名', '訂單編號', '帳號', '原有姓名', '辨識姓名（可修改）']
        col_w = [3, 22, 18, 20, 10, 14]
        for j, (h, w) in enumerate(zip(cols, col_w)):
            tk.Label(fr, text=h, bg=HDR_BG, fg=JADE, font=(FONT, 9, 'bold'),
                     width=w, anchor='w', padx=4, pady=3).grid(
                row=0, column=j, sticky='ew', padx=1, pady=(0, 2))

        check_vars = []
        name_vars  = []

        for i, item in enumerate(pending, 1):
            row_bg = WHITE if i % 2 == 0 else '#F8FBF8'
            existing = item['existing_name']
            warn     = existing and existing != item['ocr_name']

            chk_var  = tk.BooleanVar(value=True)
            name_var = tk.StringVar(value=item['ocr_name'])
            check_vars.append(chk_var)
            name_vars.append(name_var)

            tk.Checkbutton(fr, variable=chk_var, bg=row_bg).grid(
                row=i, column=0, padx=4, pady=1)

            vals = [item['fname'][:24], item['order_id'],
                    item['acct'], existing or '（無）']
            fg_list = [TEXT, GRAY, TEXT, RED if warn else GRAY]
            for j, (v, fg) in enumerate(zip(vals, fg_list), 1):
                tk.Label(fr, text=v, bg=row_bg, fg=fg, font=(FONT, 9),
                         anchor='w', padx=4, pady=2).grid(
                    row=i, column=j, sticky='ew', padx=1)

            name_entry = tk.Entry(fr, textvariable=name_var, font=(FONT, 10),
                                  relief='solid', bd=1,
                                  bg='#FFFDE7' if warn else '#F0FFF4')
            name_entry.grid(row=i, column=5, sticky='ew', padx=6, pady=2)

        fr.columnconfigure(5, weight=1)

        # ── Buttons ───────────────────────────────────────────────────────
        result = [None]

        def do_confirm():
            out = []
            for i, item in enumerate(pending):
                if not check_vars[i].get():
                    continue
                name = name_vars[i].get().strip()
                if not name:
                    continue
                out.append({**item, 'final_name': name})
            result[0] = out
            dlg.destroy()

        bf = tk.Frame(dlg, bg=WHITE)
        bf.pack(pady=10)
        self._btn(bf, '✓ 確認寫入', JADE, do_confirm).pack(side='left', padx=8)
        self._btn(bf, '✕ 全部取消', '#888888', dlg.destroy).pack(side='left', padx=8)

        dlg.wait_window()
        return result[0]

    def _dlg_new_customer(self):
        dlg = tk.Toplevel(self)
        dlg.title('新增客戶')
        dlg.geometry('380x340')
        dlg.resizable(False, False)
        dlg.configure(bg=WHITE)
        dlg.grab_set()
        dlg.transient(self)
        tk.Label(dlg, text='新增客戶', bg=WHITE, fg=JADE,
                 font=(FONT, 13, 'bold')).pack(pady=(16, 10))
        form = tk.Frame(dlg, bg=WHITE)
        form.pack(padx=24, fill='x')
        defs = [('蝦皮帳號 *', 'shopee_account'), ('真實姓名', 'real_name'),
                ('LINE 帳號',  'line_account'),   ('生　　日', 'birthday')]
        dvars = {}
        for i, (lbl, key) in enumerate(defs):
            tk.Label(form, text=lbl, bg=WHITE, fg=TEXT,
                     font=(FONT, 10), width=10, anchor='e').grid(
                row=i, column=0, sticky='e', padx=(0, 8), pady=5)
            v = tk.StringVar()
            dvars[key] = v
            tk.Entry(form, textvariable=v, font=(FONT, 10), width=22,
                     relief='solid', bd=1).grid(row=i, column=1, sticky='w')

        def submit():
            acct = dvars['shopee_account'].get().strip()
            if not acct:
                messagebox.showwarning('必填', '蝦皮帳號為必填欄位', parent=dlg)
                return
            conn = get_db()
            try:
                conn.execute(
                    'INSERT INTO customers (shopee_account,real_name,line_account,birthday) '
                    'VALUES (?,?,?,?)',
                    (acct, dvars['real_name'].get().strip(),
                     dvars['line_account'].get().strip(),
                     dvars['birthday'].get().strip()))
                conn.commit()
                dlg.destroy()
                self._refresh_list()
                if acct in self.tree.get_children():
                    self.tree.selection_set(acct)
                    self.tree.see(acct)
                self._current = acct
                self._show_detail(acct)
                self.status_var.set(f'已新增客戶：{acct}')
            except sqlite3.IntegrityError:
                messagebox.showerror('錯誤', '此蝦皮帳號已存在', parent=dlg)
            finally:
                conn.close()

        bf = tk.Frame(dlg, bg=WHITE)
        bf.pack(pady=16)
        self._btn(bf, '✓ 新增', JADE, submit).pack(side='left', padx=8)
        self._btn(bf, '取消',   GRAY, dlg.destroy).pack(side='left', padx=8)

    # ── Helper ──────────────────────────────────────────────────────────────

    @staticmethod
    def _btn(parent, text, color, cmd, small=False):
        padx = 6 if small else 12
        pady = 2 if small else 4
        fs   = 9 if small else 10
        return tk.Button(parent, text=text, bg=color, fg=WHITE,
                         font=(FONT, fs), relief='flat', cursor='hand2',
                         padx=padx, pady=pady, command=cmd,
                         activebackground=JADE_DARK, activeforeground=WHITE)


# ── Entry Point ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    App().mainloop()
